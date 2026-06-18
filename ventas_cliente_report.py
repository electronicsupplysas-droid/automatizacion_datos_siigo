#!/usr/bin/env python3
"""
Genera el reporte 'Ventas por cliente' idéntico al exportado por Siigo.

Columnas (12):
  Identificación | Sucursal | Cliente | Número de comprobantes |
  Valor bruto | Descuentos por item | Subtotal |
  Impuesto cargo | Impuesto retención |
  Cargo en totales | Descuento en totales | Total

Uso:
    python3 ventas_cliente_report.py --from-date 2026-01-01 --to-date 2026-06-30
    python3 ventas_cliente_report.py --from-date 2026-01-01 --to-date 2026-06-30 \
        --reference "Ventas por cliente.xlsx"
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from siigo_core import (
    build_siigo_config,
    customer_display_name,
    fetch_credit_notes,
    fetch_customer_details,
    fetch_paginated_results,
    invoice_exchange_rate,
    money_to_float,
    safe_text,
    to_decimal,
)

# ── Constantes ────────────────────────────────────────────────────────────────

MONEY = Decimal("0.01")

_NIT_WEIGHTS = [3, 7, 13, 17, 19, 23, 29, 37, 41, 43, 47, 53, 59, 67, 71]

# Tipos que aumentan el total (cargo al cliente)
_CARGO_TYPES = frozenset({"IVA", "Iva", "INC", "ImpConsumo"})

# Tipos que reducen el total (retención aplicada al cliente)
_RETENTION_TYPES = frozenset({"Retefuente", "ReteFuente", "ReteIVA", "ReteICA"})

# Autorretencion: impuesto del vendedor, NO afecta el total del cliente
_SKIP_TYPES = frozenset({"Autorretencion", "Autoretencion"})

COLUMNS = [
    "Identificación",
    "Sucursal",
    "Cliente",
    "Número de comprobantes",
    "Valor bruto",
    "Descuentos por item",
    "Subtotal",
    "Impuesto cargo",
    "Impuesto retención",
    "Cargo en totales",
    "Descuento en totales",
    "Total",
]

NUMERIC_COLS = {
    "Número de comprobantes",
    "Valor bruto",
    "Descuentos por item",
    "Subtotal",
    "Impuesto cargo",
    "Impuesto retención",
    "Cargo en totales",
    "Descuento en totales",
    "Total",
}


# ── Check digit NIT colombiano ────────────────────────────────────────────────

def _nit_check_digit(nit: str) -> str:
    """Dígito de verificación NIT (algoritmo DIAN)."""
    digits = nit.strip().replace("-", "").replace(" ", "")
    if not digits.isdigit():
        return ""
    total = sum(
        int(d) * _NIT_WEIGHTS[i]
        for i, d in enumerate(reversed(digits))
        if i < len(_NIT_WEIGHTS)
    )
    rem = total % 11
    return str(rem) if rem <= 1 else str(11 - rem)


def format_nit(nit: str) -> str:
    if not nit:
        return nit
    clean = nit.strip().replace("-", "").replace(" ", "")
    if not clean.isdigit():
        return nit
    return f"{clean}-{_nit_check_digit(clean)}"


# ── Desglose por documento ────────────────────────────────────────────────────

def _doc_breakdown(doc: dict[str, Any]) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """
    Calcula (valor_bruto, descuentos, impuesto_cargo, impuesto_retención) en COP
    a partir de los ítems y retenciones de un documento (factura o nota crédito).
    """
    fx = invoice_exchange_rate(doc)
    vb = disc = cargo = ret = Decimal("0")

    for item in (doc.get("items") or []):
        if not isinstance(item, dict):
            continue
        qty = to_decimal(item.get("quantity", 0))
        price = to_decimal(item.get("price", 0))
        vb += qty * price * fx

        d_obj = item.get("discount")
        if isinstance(d_obj, dict):
            disc += to_decimal(d_obj.get("value", 0)) * fx

        for tax in (item.get("taxes") or []):
            if not isinstance(tax, dict):
                continue
            t = str(tax.get("type", ""))
            v = to_decimal(tax.get("value", 0)) * fx
            if t in _CARGO_TYPES:
                cargo += v
            elif t in _RETENTION_TYPES:
                ret += v
            # _SKIP_TYPES → ignorado

    # Retenciones a nivel de cabecera (Autorretencion → ignorado, otros → retención)
    for r in (doc.get("retentions") or []):
        if not isinstance(r, dict):
            continue
        t = str(r.get("type", ""))
        if t in _RETENTION_TYPES:
            ret += to_decimal(r.get("value", 0)) * fx

    return vb, disc, cargo, ret


# ── Construcción del reporte ──────────────────────────────────────────────────

def build_ventas_por_cliente(
    invoices: list[dict[str, Any]],
    credit_notes: list[dict[str, Any]],
    customers_map: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Agrega facturas y notas crédito por cliente devolviendo las 12 columnas de Siigo.
    Las notas crédito restan de todos los componentes.
    """
    # Clave: customer_id (sin branch para agrupar igual que Siigo)
    agg: dict[str, dict[str, Any]] = {}

    for sign, docs in ((1, invoices), (-1, credit_notes)):
        for doc in docs:
            if doc.get("annulled"):
                continue
            cust = doc.get("customer") or {}
            cid = safe_text(cust.get("id")) or "unknown"
            nit = safe_text(cust.get("identification"))
            key = cid

            cname = customer_display_name(customers_map.get(cid), nit or cid)
            vb, disc, cargo, ret = _doc_breakdown(doc)

            entry = agg.setdefault(key, {
                "nit": nit,
                "cliente": cname,
                "num": 0,
                "vb": Decimal("0"),
                "disc": Decimal("0"),
                "cargo": Decimal("0"),
                "ret": Decimal("0"),
                "total_api": Decimal("0"),
            })
            entry["num"] += 1
            entry["vb"] += sign * vb
            entry["disc"] += sign * disc
            entry["cargo"] += sign * cargo
            entry["ret"] += sign * ret
            fx = invoice_exchange_rate(doc)
            entry["total_api"] += sign * to_decimal(doc.get("total", 0)) * fx

    rows = []
    for entry in sorted(agg.values(), key=lambda e: e["cliente"]):
        subtotal = entry["vb"] - entry["disc"]
        total = subtotal + entry["cargo"] - entry["ret"]
        rows.append({
            "Identificación": format_nit(entry["nit"]),
            "Sucursal": None,
            "Cliente": entry["cliente"],
            "Número de comprobantes": entry["num"],
            "Valor bruto": float(entry["vb"].quantize(MONEY, rounding=ROUND_HALF_UP)),
            "Descuentos por item": float(entry["disc"].quantize(MONEY, rounding=ROUND_HALF_UP)),
            "Subtotal": float(subtotal.quantize(MONEY, rounding=ROUND_HALF_UP)),
            "Impuesto cargo": float(entry["cargo"].quantize(MONEY, rounding=ROUND_HALF_UP)),
            "Impuesto retención": float(entry["ret"].quantize(MONEY, rounding=ROUND_HALF_UP)),
            "Cargo en totales": 0.0,
            "Descuento en totales": 0.0,
            "Total": float(entry["total_api"].quantize(MONEY, rounding=ROUND_HALF_UP)),
            "_total_api": float(entry["total_api"].quantize(MONEY, rounding=ROUND_HALF_UP)),
        })
    return rows


# ── Excel ─────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_TOTAL_FILL = PatternFill("solid", fgColor="D6DCE4")
_WHITE = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_NUM_FMT = '#,##0.00'


def _write_cell(ws: Any, row: int, col: int, value: Any, font: Any = None,
                fill: Any = None, number_format: str | None = None,
                alignment: Any = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment


def write_ventas_por_cliente_xlsx(
    output_path: Path,
    rows: list[dict[str, Any]],
    from_date: date,
    to_date: date,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas por cliente"

    # ── Filas de encabezado del informe (igual a Siigo) ──────────────────────
    ws["A1"] = "Ventas por cliente"
    ws["A1"].font = Font(bold=True, size=12)

    from_str = from_date.strftime("%-d de %B %Y").replace(
        from_date.strftime("%B"),
        _MONTH_ES[from_date.month],
    )
    to_str = to_date.strftime("%-d de %B %Y").replace(
        to_date.strftime("%B"),
        _MONTH_ES[to_date.month],
    )
    ws["A2"] = f"De {from_str} a {to_str} "

    # ── Cabeceras de columnas (fila 4) ───────────────────────────────────────
    HEADER_ROW = 4
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        _write_cell(ws, HEADER_ROW, col_idx, col_name,
                    font=_WHITE, fill=_HEADER_FILL,
                    alignment=Alignment(horizontal="center", wrap_text=True))

    # ── Filas de datos ────────────────────────────────────────────────────────
    data_rows = [r for r in rows]
    for row_idx, row_data in enumerate(data_rows, start=HEADER_ROW + 1):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val = row_data.get(col_name)
            fmt = _NUM_FMT if col_name in NUMERIC_COLS else None
            _write_cell(ws, row_idx, col_idx, val, number_format=fmt)

    # ── Fila Total general ────────────────────────────────────────────────────
    total_row = HEADER_ROW + len(data_rows) + 1
    _write_cell(ws, total_row, 1, "Total general", font=_BOLD, fill=_TOTAL_FILL)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        if col_idx == 1:
            continue
        if col_name in NUMERIC_COLS:
            total = sum(r.get(col_name, 0) or 0 for r in data_rows)
            _write_cell(ws, total_row, col_idx, round(total, 2),
                        font=_BOLD, fill=_TOTAL_FILL, number_format=_NUM_FMT)
        else:
            _write_cell(ws, total_row, col_idx, None, fill=_TOTAL_FILL)

    # ── Anchos de columna ─────────────────────────────────────────────────────
    col_widths = {
        "Identificación": 16,
        "Sucursal": 10,
        "Cliente": 50,
        "Número de comprobantes": 14,
        "Valor bruto": 18,
        "Descuentos por item": 18,
        "Subtotal": 18,
        "Impuesto cargo": 18,
        "Impuesto retención": 18,
        "Cargo en totales": 16,
        "Descuento en totales": 18,
        "Total": 18,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 15)

    ws.row_dimensions[HEADER_ROW].height = 30

    wb.save(output_path)
    print(f"  Reporte guardado: {output_path}")
    return output_path


_MONTH_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


# ── Comparación con referencia Siigo ─────────────────────────────────────────

def compare_with_reference(
    our_rows: list[dict[str, Any]],
    reference_path: Path,
) -> bool:
    """Compara nuestro reporte con el Excel de Siigo. Retorna True si coinciden."""
    import openpyxl

    wb = openpyxl.load_workbook(str(reference_path), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Detectar fila de headers
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if row and row[0] == "Identificación":
            header_row_idx = i
            break
    if header_row_idx is None:
        print("  ⚠  No se encontró la fila de encabezados en la referencia.")
        return False

    # Leer filas de referencia (solo filas con NIT válido: dígitos con dígito verificación)
    import re as _re
    _NIT_PATTERN = _re.compile(r'^\d{5,12}-\d$')

    ref_by_nit: dict[str, dict[str, float]] = {}
    for row in all_rows[header_row_idx + 1:]:
        if not row or not row[0]:
            continue
        id_val = str(row[0]).strip()
        if not _NIT_PATTERN.match(id_val):
            continue  # Ignora "Total general", pie de página, etc.
        # Columnas: 0=Id, 1=Sucursal, 2=Cliente, 3=N°, 4=VB, 5=Disc, 6=Sub, 7=Cargo, 8=Ret, 9=CT, 10=DT, 11=Total
        try:
            ref_by_nit[id_val] = {
                "Valor bruto": float(row[4] or 0),
                "Descuentos por item": float(row[5] or 0),
                "Subtotal": float(row[6] or 0),
                "Impuesto cargo": float(row[7] or 0),
                "Impuesto retención": float(row[8] or 0),
                "Total": float(row[11] or 0),
                "N°": int(row[3] or 0),
                "Cliente": str(row[2] or ""),
            }
        except (TypeError, ValueError):
            continue

    # Nuestro reporte indexado por NIT
    our_by_nit: dict[str, dict[str, Any]] = {r["Identificación"]: r for r in our_rows}

    print(f"\n{'='*80}")
    print(f"  COMPARACIÓN CON REFERENCIA SIIGO ({reference_path.name})")
    print(f"{'='*80}")

    tolerance = Decimal("0.05")
    all_match = True
    mismatches: list[str] = []
    missing_in_ours: list[str] = []
    missing_in_ref: list[str] = []

    for nit, ref in sorted(ref_by_nit.items()):
        our = our_by_nit.get(nit)
        if our is None:
            missing_in_ours.append(f"  FALTA EN NUESTRO REPORTE: {nit} ({ref['Cliente']})")
            all_match = False
            continue
        diffs = []
        for col in ("Valor bruto", "Descuentos por item", "Subtotal",
                    "Impuesto cargo", "Impuesto retención", "Total"):
            ref_val = round(ref[col], 2)
            our_val = round(our.get(col, 0), 2)
            diff = abs(ref_val - our_val)
            if diff > float(tolerance):
                diffs.append(f"    {col}: Siigo={ref_val:,.2f}  Nuestro={our_val:,.2f}  Δ={diff:,.2f}")
        if ref["N°"] != our.get("Número de comprobantes", 0):
            diffs.append(f"    N° comprobantes: Siigo={ref['N°']}  Nuestro={our.get('Número de comprobantes',0)}")

        if diffs:
            mismatches.append(f"  ❌ {nit} | {ref['Cliente']}")
            mismatches.extend(diffs)
            all_match = False
        else:
            print(f"  ✓  {nit} | {ref['Cliente']}")

    for nit in our_by_nit:
        if nit not in ref_by_nit:
            missing_in_ref.append(f"  EXTRA EN NUESTRO REPORTE (no en Siigo): {nit} ({our_by_nit[nit]['Cliente']})")

    if missing_in_ours:
        print("\n  --- FALTANTES EN NUESTRO REPORTE ---")
        for m in missing_in_ours:
            print(m)
    if missing_in_ref:
        print("\n  --- EXTRA EN NUESTRO REPORTE (nuevos desde la referencia) ---")
        for m in missing_in_ref:
            print(m)
    if mismatches:
        print("\n  --- DIFERENCIAS ---")
        for m in mismatches:
            print(m)
        print()
        print("  NOTAS SOBRE DIFERENCIAS:")
        print("  · Si hay diferencias en N° o valores, puede ser porque la referencia Siigo fue")
        print("    generada antes de que esas facturas se procesaran (diferencia de horario).")
        print("  · QUICENO (y cualquier cliente con documentos DF): Siigo incluye documentos tipo")
        print("    'Débito Facturación' (prefijo DF) que NO están disponibles en la API pública")
        print("    de Siigo (v1/invoices, v1/credit-notes ni ningún otro endpoint conocido).")
        print("    Para igualar exactamente, se requeriría acceso al endpoint DF que Siigo no expone.")

    # Totales globales
    print("\n  --- TOTALES GLOBALES ---")
    for col in ("Valor bruto", "Descuentos por item", "Subtotal",
                "Impuesto cargo", "Impuesto retención", "Total"):
        ref_total = round(sum(v[col] for v in ref_by_nit.values()), 2)
        our_total = round(sum(r.get(col, 0) for r in our_rows), 2)
        status = "✓" if abs(ref_total - our_total) <= float(tolerance) else "❌"
        print(f"  {status} {col}: Siigo={ref_total:>22,.2f}  Nuestro={our_total:>22,.2f}  Δ={abs(ref_total-our_total):,.2f}")

    print(f"{'='*80}\n")
    return all_match


# ── Fetch ─────────────────────────────────────────────────────────────────────

def fetch_invoices(config: Any, from_date: date, to_date: date, timeout: int = 60) -> list[dict]:
    fetched = fetch_paginated_results(
        config,
        "v1/invoices",
        {
            "page": "1",
            "page_size": "100",
            "date_start": f"{from_date.isoformat()}T00:00:00Z",
            "date_end": f"{to_date.isoformat()}T23:59:59Z",
        },
        timeout=timeout,
    )
    from_str = from_date.isoformat()
    to_str = to_date.isoformat()
    return [
        r for r in fetched["results"]
        if isinstance(r, dict)
        and from_str <= (safe_text(r.get("date")) or "")[:10] <= to_str
    ]


# ── CLI ───────────────────────────────────────────────────────────────────────

def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Reporte Ventas por Cliente (idéntico a Siigo)")
    p.add_argument("--from-date", required=True, help="Fecha inicial YYYY-MM-DD")
    p.add_argument("--to-date", required=True, help="Fecha final YYYY-MM-DD")
    p.add_argument("--output", default="output/informes/ventas_por_cliente.xlsx",
                   help="Ruta del Excel de salida")
    p.add_argument("--reference", default=None,
                   help="Ruta al Excel de Siigo para comparar")
    p.add_argument("--timeout", type=int, default=60)
    return p.parse_args()


def main() -> int:
    args = _parse_args()
    from_date = date.fromisoformat(args.from_date)
    to_date = date.fromisoformat(args.to_date)
    output_path = Path(args.output)

    config = build_siigo_config()

    print(f"Descargando facturas {from_date} → {to_date}…")
    invoices = fetch_invoices(config, from_date, to_date, timeout=args.timeout)
    print(f"  {len(invoices)} facturas")

    print("Descargando notas crédito…")
    credit_notes = fetch_credit_notes(config, from_date, to_date, timeout=args.timeout)
    print(f"  {len(credit_notes)} notas crédito")

    # Recopilar IDs de clientes únicos (de facturas + NCs)
    customer_hints: dict[str, dict[str, str]] = {}
    for doc in invoices + credit_notes:
        cust = doc.get("customer")
        if not isinstance(cust, dict):
            continue
        cid = safe_text(cust.get("id"))
        if cid:
            customer_hints[cid] = {"identification": safe_text(cust.get("identification"))}

    print(f"Descargando detalles de {len(customer_hints)} clientes…")
    customers_map, failures = fetch_customer_details(config, customer_hints, timeout=args.timeout)
    if failures:
        print(f"  ⚠  {len(failures)} clientes sin detalle")

    print("Calculando reporte…")
    rows = build_ventas_por_cliente(invoices, credit_notes, customers_map)
    print(f"  {len(rows)} clientes")

    write_ventas_por_cliente_xlsx(output_path, rows, from_date, to_date)

    if args.reference:
        ref_path = Path(args.reference)
        if ref_path.exists():
            compare_with_reference(rows, ref_path)
        else:
            print(f"  ⚠  Archivo de referencia no encontrado: {ref_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
