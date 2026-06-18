#!/usr/bin/env python3
"""
cartera_report.py — Cuentas por cobrar detallada por documento.

Replica el informe de Siigo 'Cuentas por cobrar detallada por documento':
  • Una fila por factura con saldo pendiente (balance > 0)
  • Antigüedad calculada desde la fecha de vencimiento real (payments[].due_date),
    NO desde la fecha de emisión (bug del reporte anterior)
  • Centro de costo y ciudad del cliente incluidos

Columnas (15):
  Identificación | Cliente | Sucursal | Documento | Fecha vencimiento |
  Centro de costo | Cobrador | Vendedor | Ciudad |
  Vencido 1 a 30 | Vencido 31 a 60 | Vencido 61 a 90 | Vencido más de 91 |
  Saldo por vencer | Saldo a favor | Total cartera

Uso:
    python3 cartera_report.py                                  # últimos 36 meses
    python3 cartera_report.py --from-date 2024-01-01
    python3 cartera_report.py --from-date 2024-01-01 --to-date 2026-06-30
    python3 cartera_report.py --output output/informes/cartera_202606.xlsx
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from siigo_core import (
    SiigoApiError,
    build_siigo_config,
    customer_display_name,
    fetch_customer_details,
    fetch_paginated_results,
    fetch_users_map,
    invoice_exchange_rate,
    normalize_path,
    obtain_token,
    readonly_get_url,
    safe_text,
    build_get_url,
    to_decimal,
    user_display_name,
)

# ── Constantes ────────────────────────────────────────────────────────────────

MONEY = Decimal("0.01")

COLUMNS = [
    "Identificación",
    "Cliente",
    "Sucursal",
    "Documento",
    "Fecha vencimiento",
    "Centro de costo",
    "Cobrador",
    "Vendedor",
    "Ciudad",
    "Vencido 1 a 30",
    "Vencido 31 a 60",
    "Vencido 61 a 90",
    "Vencido más de 91",
    "Saldo por vencer",
    "Saldo a favor",
    "Total cartera",
]

MONEY_COLS = {
    "Vencido 1 a 30", "Vencido 31 a 60", "Vencido 61 a 90",
    "Vencido más de 91", "Saldo por vencer", "Saldo a favor", "Total cartera",
}

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_TOTAL_FILL   = PatternFill("solid", fgColor="D6DCE4")
_WHITE_BOLD   = Font(color="FFFFFF", bold=True)
_BOLD         = Font(bold=True)
_NUM_FMT      = '#,##0.00'


# ── Helpers ────────────────────────────────────────────────────────────────────

def _due_date_from_invoice(invoice: dict[str, Any]) -> str:
    """
    Extrae la fecha de vencimiento real desde payments[].due_date.
    El campo invoice.due_date siempre viene vacío en la API de Siigo.
    """
    payments = invoice.get("payments") or []
    dates = [
        p["due_date"][:10]
        for p in payments
        if isinstance(p, dict) and p.get("due_date")
    ]
    return max(dates) if dates else ""


def _aging_buckets(
    balance: Decimal,
    due_date_str: str,
    today: date,
) -> tuple[Decimal, Decimal, Decimal, Decimal, Decimal, Decimal]:
    """
    Distribuye el saldo en los 6 buckets de antigüedad:
      (por_vencer, v1_30, v31_60, v61_90, v91_plus, a_favor)

    La antigüedad se mide en días desde la fecha de vencimiento, no desde la
    fecha de emisión. Esto es lo que hace Siigo en su informe de cartera.
    """
    z = Decimal("0")

    # Saldo a favor: el cliente pagó de más (balance negativo no debería ocurrir
    # con el filtro balance > 0, pero lo manejamos por robustez)
    if balance < 0:
        return z, z, z, z, z, abs(balance)

    if not due_date_str:
        # Sin fecha de vencimiento → tratar como vencido >91 (posición conservadora)
        return z, z, z, z, balance, z

    try:
        due = date.fromisoformat(due_date_str[:10])
    except ValueError:
        return z, z, z, z, balance, z

    days = (today - due).days  # positivo = vencido, negativo = por vencer

    if days <= 0:
        return balance, z, z, z, z, z  # no vencida aún
    elif days <= 30:
        return z, balance, z, z, z, z
    elif days <= 60:
        return z, z, balance, z, z, z
    elif days <= 90:
        return z, z, z, balance, z, z
    else:
        return z, z, z, z, balance, z  # "más de 91" días


def _city_from_customer(customer_body: dict[str, Any] | None) -> str:
    if not isinstance(customer_body, dict):
        return ""
    address = customer_body.get("address")
    if not isinstance(address, dict):
        return ""
    city = address.get("city")
    if not isinstance(city, dict):
        return city if isinstance(city, str) else ""
    return safe_text(city.get("city_name"))


# ── Fetch de catálogos auxiliares ─────────────────────────────────────────────

def fetch_cost_centers(config: Any, timeout: int = 30) -> dict[int, str]:
    """
    Descarga el catálogo de centros de costo y devuelve {id: nombre}.
    El endpoint retorna una lista directa (no paginada).
    """
    token, token_type = obtain_token(config, timeout=timeout)
    url = build_get_url(config, "v1/cost-centers", {})
    _, body = readonly_get_url(config, url, timeout=timeout, token=token, token_type=token_type)

    if not isinstance(body, list):
        return {}

    return {
        item["id"]: safe_text(item.get("name"))
        for item in body
        if isinstance(item, dict) and item.get("id") is not None
    }


def fetch_cartera_invoices(
    config: Any,
    from_date: date,
    to_date: date,
    timeout: int = 90,
) -> list[dict[str, Any]]:
    """
    Descarga todas las facturas en el rango de fecha de emisión dado.
    Aplica filtros client-side: no anuladas, balance > 0.
    """
    fetched = fetch_paginated_results(
        config,
        "v1/invoices",
        {
            "page": "1",
            "page_size": "100",
            "date_start": f"{from_date.isoformat()}T00:00:00Z",
            "date_end": f"{to_date.isoformat()}T23:59:59Z",
        },
        timeout=timeout,
    )

    result = []
    for inv in fetched["results"]:
        if not isinstance(inv, dict):
            continue
        if inv.get("annulled"):
            continue
        fx = invoice_exchange_rate(inv)
        balance = to_decimal(inv.get("balance", 0)) * fx
        if balance > 0:
            result.append(inv)
    return result


# ── Construcción de filas ─────────────────────────────────────────────────────

def build_cartera_rows(
    invoices: list[dict[str, Any]],
    cost_centers_map: dict[int, str],
    users_map: dict[str, dict[str, Any]],
    customers_map: dict[str, dict[str, Any]],
    as_of: date,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for inv in invoices:
        fx = invoice_exchange_rate(inv)
        balance = to_decimal(inv.get("balance", 0)) * fx

        # Cliente
        cust = inv.get("customer") if isinstance(inv.get("customer"), dict) else {}
        cid  = safe_text(cust.get("id"))
        nit  = safe_text(cust.get("identification"))
        cust_body = customers_map.get(cid)
        cname     = customer_display_name(cust_body, nit or cid)
        city      = _city_from_customer(cust_body)

        # Vendedor
        seller_id   = safe_text(inv.get("seller"))
        seller_body = users_map.get(seller_id) if seller_id else None
        seller_name = user_display_name(seller_body, "") if seller_body else ""

        # Centro de costo
        cc_id   = inv.get("cost_center")
        cc_name = cost_centers_map.get(cc_id, "") if cc_id is not None else ""

        # Fecha de vencimiento (desde payments, no desde el raíz del documento)
        due_date_str = _due_date_from_invoice(inv)

        # Buckets de antigüedad
        por_vencer, v1, v31, v61, v91, a_favor = _aging_buckets(balance, due_date_str, as_of)

        def m(val: Decimal) -> float:
            return float(val.quantize(MONEY, rounding=ROUND_HALF_UP))

        rows.append({
            "Identificación":   nit,
            "Cliente":          cname,
            "Sucursal":         safe_text(cust.get("branch_office")) or None,
            "Documento":        safe_text(inv.get("name")),
            "Fecha vencimiento": due_date_str or None,
            "Centro de costo":  cc_name or None,
            "Cobrador":         None,  # no expuesto en API pública
            "Vendedor":         seller_name or None,
            "Ciudad":           city or None,
            "Vencido 1 a 30":  m(v1),
            "Vencido 31 a 60": m(v31),
            "Vencido 61 a 90": m(v61),
            "Vencido más de 91": m(v91),
            "Saldo por vencer": m(por_vencer),
            "Saldo a favor":    m(a_favor),
            "Total cartera":    m(balance),
        })

    # Ordenar igual que Siigo: por cliente, luego por fecha de vencimiento
    rows.sort(key=lambda r: (r["Cliente"] or "", r["Fecha vencimiento"] or ""))
    return rows


# ── Excel ─────────────────────────────────────────────────────────────────────

def write_cartera_xlsx(
    output_path: Path,
    rows: list[dict[str, Any]],
    as_of: date,
    from_date: date,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera"

    # ── Encabezado del informe ────────────────────────────────────────────────
    ws["A1"] = "Cuentas por cobrar detallada por documento"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"Al {as_of.strftime('%d/%m/%Y')}  (facturas desde {from_date.strftime('%d/%m/%Y')})"

    HEADER_ROW = 4

    # ── Cabeceras de columnas ─────────────────────────────────────────────────
    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=col)
        cell.font  = _WHITE_BOLD
        cell.fill  = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[HEADER_ROW].height = 30

    # ── Filas de datos ────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, start=HEADER_ROW + 1):
        for ci, col in enumerate(COLUMNS, 1):
            val = row.get(col)
            cell = ws.cell(row=ri, column=ci, value=val)
            if col in MONEY_COLS and isinstance(val, (int, float)):
                cell.number_format = _NUM_FMT
                cell.alignment = Alignment(horizontal="right")

    # ── Fila totales ──────────────────────────────────────────────────────────
    total_row = HEADER_ROW + len(rows) + 1
    ws.cell(total_row, 1, "Total general").font = _BOLD
    ws.cell(total_row, 1).fill = _TOTAL_FILL
    for ci, col in enumerate(COLUMNS, 1):
        if col in MONEY_COLS:
            total = round(sum(r.get(col, 0) or 0 for r in rows), 2)
            cell = ws.cell(total_row, ci, total)
            cell.font = _BOLD
            cell.fill = _TOTAL_FILL
            cell.number_format = _NUM_FMT
            cell.alignment = Alignment(horizontal="right")
        elif ci > 1:
            ws.cell(total_row, ci).fill = _TOTAL_FILL

    # ── Anchos de columna ─────────────────────────────────────────────────────
    widths = {
        "Identificación": 16,
        "Cliente": 42,
        "Sucursal": 10,
        "Documento": 14,
        "Fecha vencimiento": 16,
        "Centro de costo": 20,
        "Cobrador": 14,
        "Vendedor": 28,
        "Ciudad": 18,
        "Vencido 1 a 30": 16,
        "Vencido 31 a 60": 16,
        "Vencido 61 a 90": 16,
        "Vencido más de 91": 18,
        "Saldo por vencer": 16,
        "Saldo a favor": 14,
        "Total cartera": 18,
    }
    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)

    wb.save(output_path)
    print(f"  Reporte guardado: {output_path}")
    return output_path


# ── Comparación con referencia Siigo ─────────────────────────────────────────

def compare_with_reference(
    our_rows: list[dict[str, Any]],
    reference_path: Path,
) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(str(reference_path), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Detectar fila de headers
    header_idx = next(
        (i for i, r in enumerate(all_rows) if r and r[0] == "Identificación"), None
    )
    if header_idx is None:
        print("  ⚠  No se encontró la fila 'Identificación' en la referencia.")
        return

    # Indexar referencia por documento (nombre de factura)
    ref_by_doc: dict[str, dict] = {}
    for row in all_rows[header_idx + 1:]:
        if not row or not row[3]:  # col 3 = Documento
            continue
        doc = str(row[3]).strip()
        if not doc.startswith("FV"):
            continue
        ref_by_doc[doc] = {
            "nit": str(row[0] or ""),
            "cliente": str(row[1] or ""),
            "due_date": str(row[4] or ""),
            "v1": float(row[9] or 0),
            "v31": float(row[10] or 0),
            "v61": float(row[11] or 0),
            "v91": float(row[12] or 0),
            "por_vencer": float(row[13] or 0),
            "a_favor": float(row[14] or 0),
            "total": float(row[15] or 0),
        }

    our_by_doc = {r["Documento"]: r for r in our_rows}

    print(f"\n{'='*80}")
    print(f"  COMPARACIÓN CON REFERENCIA  ({reference_path.name})")
    print(f"{'='*80}")

    tol = 0.05
    ok = miss_ours = miss_ref = diff = 0

    for doc, ref in sorted(ref_by_doc.items()):
        our = our_by_doc.get(doc)
        if our is None:
            print(f"  FALTA en nuestro reporte: {doc} ({ref['cliente']})")
            miss_ours += 1
            continue
        errs = []
        for label, ref_val, our_key in [
            ("Total",          ref["total"],     "Total cartera"),
            ("Vencido 1-30",   ref["v1"],        "Vencido 1 a 30"),
            ("Vencido 31-60",  ref["v31"],       "Vencido 31 a 60"),
            ("Vencido 61-90",  ref["v61"],       "Vencido 61 a 90"),
            ("Vencido >91",    ref["v91"],       "Vencido más de 91"),
            ("Por vencer",     ref["por_vencer"],"Saldo por vencer"),
        ]:
            d = abs(ref_val - (our.get(our_key) or 0))
            if d > tol:
                errs.append(f"    {label}: Siigo={ref_val:,.2f}  Nuestro={our.get(our_key, 0):,.2f}  Δ={d:,.2f}")
        if errs:
            print(f"  ❌ {doc} | {ref['cliente']}")
            for e in errs:
                print(e)
            diff += 1
        else:
            ok += 1

    for doc in our_by_doc:
        if doc not in ref_by_doc:
            miss_ref += 1

    print(f"\n  ✓ {ok} documentos coinciden")
    if diff:
        print(f"  ❌ {diff} con diferencias")
    if miss_ours:
        print(f"  ⚠  {miss_ours} en Siigo pero no en nuestro reporte")
    if miss_ref:
        print(f"  +  {miss_ref} en nuestro reporte pero no en la referencia (nuevos)")

    # Totales globales
    print("\n  --- TOTALES GLOBALES ---")
    totals_ref = {k: sum(v[k] for v in ref_by_doc.values())
                  for k in ("v1", "v31", "v61", "v91", "por_vencer", "total")}
    totals_our = {
        "v1":         sum(r.get("Vencido 1 a 30", 0) or 0 for r in our_rows),
        "v31":        sum(r.get("Vencido 31 a 60", 0) or 0 for r in our_rows),
        "v61":        sum(r.get("Vencido 61 a 90", 0) or 0 for r in our_rows),
        "v91":        sum(r.get("Vencido más de 91", 0) or 0 for r in our_rows),
        "por_vencer": sum(r.get("Saldo por vencer", 0) or 0 for r in our_rows),
        "total":      sum(r.get("Total cartera", 0) or 0 for r in our_rows),
    }
    labels = {
        "v1": "Vencido 1-30", "v31": "Vencido 31-60",
        "v61": "Vencido 61-90", "v91": "Vencido >91",
        "por_vencer": "Saldo por vencer", "total": "Total cartera",
    }
    for k, label in labels.items():
        d = abs(totals_ref[k] - totals_our[k])
        icon = "✓" if d <= tol else "❌"
        print(f"  {icon} {label:<20} Siigo={totals_ref[k]:>22,.2f}  "
              f"Nuestro={totals_our[k]:>22,.2f}  Δ={d:,.2f}")
    print(f"{'='*80}\n")


# ── CLI ───────────────────────────────────────────────────────────────────────

def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Cartera — cuentas por cobrar por documento")
    p.add_argument(
        "--from-date",
        default=(date.today() - timedelta(days=365 * 3)).isoformat(),
        help="Fecha inicial de emisión (default: hace 36 meses)",
    )
    p.add_argument(
        "--to-date",
        default=date.today().isoformat(),
        help="Fecha final de emisión (default: hoy)",
    )
    p.add_argument(
        "--as-of",
        default=date.today().isoformat(),
        help="Fecha de corte para calcular antigüedad (default: hoy)",
    )
    p.add_argument(
        "--output",
        default="output/informes/cartera.xlsx",
        help="Ruta del archivo de salida",
    )
    p.add_argument(
        "--reference",
        default=None,
        help="Excel de Siigo para comparar (opcional)",
    )
    p.add_argument("--timeout", type=int, default=90)
    return p.parse_args()


def main() -> int:
    args = _parse_args()

    try:
        from_date = date.fromisoformat(args.from_date)
        to_date   = date.fromisoformat(args.to_date)
        as_of     = date.fromisoformat(args.as_of)
    except ValueError as exc:
        print(f"ERROR: fecha inválida — {exc}", file=sys.stderr)
        return 1

    output_path = Path(args.output)
    config = build_siigo_config()

    print(f"Descargando centros de costo…")
    cost_centers_map = fetch_cost_centers(config, timeout=args.timeout)
    print(f"  {len(cost_centers_map)} centros de costo.")

    print(f"Descargando vendedores…")
    users_map, _ = fetch_users_map(config, timeout=args.timeout)
    print(f"  {len(users_map)} vendedores.")

    print(f"Descargando facturas con saldo pendiente ({from_date} → {to_date})…")
    invoices = fetch_cartera_invoices(config, from_date, to_date, timeout=args.timeout)
    print(f"  {len(invoices)} facturas con balance > 0.")

    # Recopilar clientes únicos
    customer_hints: dict[str, dict[str, str]] = {}
    for inv in invoices:
        cust = inv.get("customer")
        if not isinstance(cust, dict):
            continue
        cid = safe_text(cust.get("id"))
        if cid:
            customer_hints[cid] = {"identification": safe_text(cust.get("identification"))}

    print(f"Descargando detalles de {len(customer_hints)} clientes (para ciudad)…")
    customers_map, failures = fetch_customer_details(config, customer_hints, timeout=args.timeout)
    if failures:
        print(f"  ⚠  {len(failures)} clientes sin detalle.", file=sys.stderr)

    print(f"Calculando cartera al {as_of}…")
    rows = build_cartera_rows(invoices, cost_centers_map, users_map, customers_map, as_of)
    print(f"  {len(rows)} documentos en cartera.")

    write_cartera_xlsx(output_path, rows, as_of, from_date)

    # Resumen por bucket
    total         = sum(r["Total cartera"] for r in rows)
    por_vencer    = sum(r["Saldo por vencer"] for r in rows)
    v1            = sum(r["Vencido 1 a 30"] for r in rows)
    v31           = sum(r["Vencido 31 a 60"] for r in rows)
    v61           = sum(r["Vencido 61 a 90"] for r in rows)
    v91           = sum(r["Vencido más de 91"] for r in rows)

    print(f"\n{'─'*55}")
    print(f"  Total cartera al {as_of}  :  $ {total:>20,.2f}")
    print(f"  Saldo por vencer          :  $ {por_vencer:>20,.2f}  ({100*por_vencer/total:.1f}%)" if total else "")
    print(f"  Vencido   1 -  30 días    :  $ {v1:>20,.2f}  ({100*v1/total:.1f}%)" if total else "")
    print(f"  Vencido  31 -  60 días    :  $ {v31:>20,.2f}  ({100*v31/total:.1f}%)" if total else "")
    print(f"  Vencido  61 -  90 días    :  $ {v61:>20,.2f}  ({100*v61/total:.1f}%)" if total else "")
    print(f"  Vencido  > 90 días        :  $ {v91:>20,.2f}  ({100*v91/total:.1f}%)" if total else "")
    print(f"{'─'*55}\n")

    if args.reference:
        ref = Path(args.reference)
        if ref.exists():
            compare_with_reference(rows, ref)
        else:
            print(f"  ⚠  Referencia no encontrada: {ref}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
