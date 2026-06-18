#!/usr/bin/env python3
"""
Reporte de comisiones por vendedor — 3% sobre saldo neto de facturas.

Saldo neto = total factura − notas crédito vinculadas.
Mes de comisión = mes del recibo de caja (fecha real de cobro).
Solo se liquidan facturas que tienen al menos un recibo de caja — las facturas sin pagar no generan comisión.

Fuentes:
  v1/invoices      — facturas del período
  v1/credit-notes  — notas crédito (para calcular el saldo neto)
  v1/vouchers      — recibos de caja (para la fecha real de cobro)

Uso:
    python3 comisiones_report.py --from-date 2026-01-01 --to-date 2026-05-20 \
        --output output/comisiones_2026.xlsx
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from siigo_core import (
    SiigoApiError,
    build_siigo_config,
    fetch_paginated_results,
    fetch_users_map,
    fetch_customer_details,
    fetch_credit_notes,
    invoice_exchange_rate,
    money_to_float,
    safe_text,
    to_decimal,
    user_display_name,
    customer_display_name,
)

COMMISSION_RATE = Decimal("0.03")


# ── Helpers de fecha ───────────────────────────────────────────────────────────

def parse_iso_date(value: str, label: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise SiigoApiError(f"{label} debe tener formato YYYY-MM-DD") from exc


def rfc3339(d: date, end_of_day: bool = False) -> str:
    suffix = "T23:59:59Z" if end_of_day else "T00:00:00Z"
    return f"{d.isoformat()}{suffix}"




# ── Estilos Excel ──────────────────────────────────────────────────────────────

COLOR_HEADER = "1F4E78"
COLOR_SUB    = "2E75B6"
COLOR_TOTAL  = "D6E4F0"
WHITE        = "FFFFFF"


def _side() -> Side:
    return Side(style="thin", color="CCCCCC")


def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _hcell(cell: Any, text: str, color: str = COLOR_HEADER) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _dcell(cell: Any, value: Any, align: str = "left") -> None:
    cell.value = value
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()


def _mcell(cell: Any, value: float) -> None:
    cell.value = value
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = _border()


def _tcell(cell: Any, value: Any, is_money: bool = False) -> None:
    cell.value = value
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL)
    cell.border = _border()
    if is_money:
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_table(ws: Any, rows: int, cols: int, name: str) -> None:
    if rows < 1:
        return
    ref = f"A1:{get_column_letter(cols)}{rows + 1}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(tbl)


# ── Lógica de negocio ──────────────────────────────────────────────────────────

def fetch_invoices_in_range(
    config: Any,
    from_date: date,
    to_date: date,
    timeout: int = 60,
) -> dict[str, dict[str, Any]]:
    """Facturas emitidas en [from_date, to_date]. Devuelve {name: inv}."""
    fetched = fetch_paginated_results(
        config,
        "v1/invoices",
        {
            "page":       "1",
            "page_size":  "100",
            "date_start": rfc3339(from_date),
            "date_end":   rfc3339(to_date, end_of_day=True),
        },
        timeout=timeout,
    )
    inv_map: dict[str, dict[str, Any]] = {}
    for inv in fetched["results"]:
        if not isinstance(inv, dict):
            continue
        name = safe_text(inv.get("name"))
        if name:
            inv_map[name] = inv
    return inv_map


def fetch_receipt_map(
    config: Any,
    from_date: date,
    to_date: date,
    timeout: int = 60,
) -> dict[str, tuple[str, str]]:
    """
    Descarga todos los recibos de caja desde from_date hasta hoy y devuelve
    {invoice_name: (rc_date, rc_name)} con el recibo más reciente por factura.
    La ventana llega hasta hoy para capturar pagos recientes de facturas antiguas.
    """
    fetch_from = from_date
    fetch_to   = date.today()

    fetched = fetch_paginated_results(
        config,
        "v1/vouchers",
        {
            "page":          "1",
            "page_size":     "100",
            "created_start": rfc3339(fetch_from),
            "created_end":   rfc3339(fetch_to, end_of_day=True),
        },
        timeout=timeout,
    )

    receipt_map: dict[str, tuple[str, str]] = {}
    for v in fetched["results"]:
        if not isinstance(v, dict):
            continue
        rc_date = safe_text(v.get("date")) or ""
        rc_name = safe_text(v.get("name")) or ""
        for item in v.get("items", []):
            due = item.get("due")
            if not isinstance(due, dict):
                continue
            prefix      = safe_text(due.get("prefix"))
            consecutive = due.get("consecutive")
            if not prefix or consecutive is None:
                continue
            inv_key = f"{prefix}-{consecutive}"
            if inv_key not in receipt_map or (rc_date and rc_date > receipt_map[inv_key][0]):
                receipt_map[inv_key] = (rc_date, rc_name)
    return receipt_map


def build_commission_report(
    config: Any,
    from_date: date,
    to_date: date,
    timeout: int = 60,
) -> dict[str, Any]:

    print("  Descargando facturas del período…")
    inv_map = fetch_invoices_in_range(config, from_date, to_date, timeout=timeout)
    print(f"  {len(inv_map)} facturas encontradas.")

    print("  Descargando notas crédito…")
    try:
        raw_ncs = fetch_credit_notes(config, from_date, to_date, timeout=timeout)
    except SiigoApiError as exc:
        print(f"  Advertencia: no se pudieron obtener notas crédito: {exc}", file=sys.stderr)
        raw_ncs = []
    print(f"  {len(raw_ncs)} notas crédito encontradas.")

    print("  Descargando recibos de caja…")
    receipt_map = fetch_receipt_map(config, from_date, to_date, timeout=timeout)
    print(f"  {len(receipt_map)} facturas con recibo asociado.")

    print("  Descargando vendedores…")
    users_map, _ = fetch_users_map(config, timeout=timeout)

    customer_hints: dict[str, dict[str, str]] = {}
    for inv in inv_map.values():
        c = inv.get("customer")
        if isinstance(c, dict):
            cid = safe_text(c.get("id"))
            if cid:
                customer_hints[cid] = {"identification": safe_text(c.get("identification"))}
    customers_map, _ = fetch_customer_details(config, customer_hints, timeout=timeout)

    # ── Mapear NCs a facturas ─────────────────────────────────────────────────
    nc_by_invoice: dict[str, list[dict]] = defaultdict(list)
    orphan_ncs: list[dict] = []

    for nc in raw_ncs:
        if nc.get("annulled"):
            continue
        inv_ref = nc.get("invoice")
        if isinstance(inv_ref, dict):
            inv_name = safe_text(inv_ref.get("name"))
            if inv_name and inv_name in inv_map:
                nc_by_invoice[inv_name].append(nc)
            else:
                orphan_ncs.append(nc)
        else:
            orphan_ncs.append(nc)

    # ── Construir filas de detalle — una por factura ──────────────────────────
    # Todas las facturas se procesan; se separan en pagadas y pendientes.
    # La base de comisión cubre el mismo universo que ventas/facturación.
    detail_rows: list[dict[str, Any]] = []
    pending_rows: list[dict[str, Any]] = []
    seller_month: dict[str, dict[str, dict[str, Any]]] = defaultdict(
        lambda: defaultdict(lambda: {
            "saldo_neto":       Decimal("0"),
            "comision_pagada":  Decimal("0"),
            "facturas_pagadas": 0,
        })
    )
    seller_total: dict[str, dict[str, Any]] = {}
    all_months: set[str] = set()

    for inv_name, inv in inv_map.items():
        inv_date_str = safe_text(inv.get("date"))
        inv_date     = inv_date_str[:10] if inv_date_str else ""
        inv_month    = inv_date[:7] if inv_date else ""

        fx        = invoice_exchange_rate(inv)
        inv_total = to_decimal(inv.get("total")) * fx

        # Moneda original de la factura
        currency_info = inv.get("currency")
        moneda = safe_text(currency_info.get("code")) if isinstance(currency_info, dict) else "COP"
        moneda = moneda or "COP"

        # Notas crédito vinculadas — cada NC usa su propia tasa de cambio
        ncs_for_inv = nc_by_invoice.get(inv_name, [])
        nc_total = sum(
            (to_decimal(nc.get("total")) * invoice_exchange_rate(nc) for nc in ncs_for_inv),
            Decimal("0"),
        )
        nc_ids = ", ".join(
            safe_text(nc.get("name")) for nc in ncs_for_inv if safe_text(nc.get("name"))
        )

        # Base de comisión: saldo neto (total - NCs).
        # Para facturas sin recibo de caja (pagadas por otro medio), se usa
        # el balance real de la API si está disponible, para no sobrestimar.
        balance_api = to_decimal(inv.get("balance", inv.get("total", 0))) * fx
        saldo_neto_base = min(inv_total, balance_api) if balance_api > 0 else inv_total
        saldo_neto   = max(saldo_neto_base - nc_total, Decimal("0"))
        commission   = saldo_neto * COMMISSION_RATE

        # Vendedor
        seller_id   = safe_text(inv.get("seller")) or "unknown"
        seller_body = users_map.get(seller_id)
        seller_name = user_display_name(
            seller_body,
            seller_id if seller_id != "unknown" else "Sin vendedor",
        )

        # Cliente
        c     = inv.get("customer") if isinstance(inv.get("customer"), dict) else {}
        cid   = safe_text(c.get("id"))
        cname = customer_display_name(
            customers_map.get(cid),
            safe_text(c.get("identification")) or cid or "Sin cliente",
        )

        # Inicializar acumulador del vendedor (cubre todas las facturas)
        if seller_id not in seller_total:
            seller_total[seller_id] = {
                "vendedor_id":         seller_id,
                "vendedor":            seller_name,
                "facturas_total":      0,
                "facturas_pagadas":    0,
                "facturas_pendientes": 0,
                "total_factura":       Decimal("0"),
                "total_nc":            Decimal("0"),
                "saldo_neto":          Decimal("0"),
                "comision_pagada":     Decimal("0"),
                "comision_pendiente":  Decimal("0"),
            }
        st = seller_total[seller_id]
        st["facturas_total"]  += 1
        st["total_factura"]   += inv_total
        st["total_nc"]        += nc_total
        st["saldo_neto"]      += saldo_neto

        rc_info = receipt_map.get(inv_name)
        if rc_info:
            # Factura pagada — comisión en el mes del recibo
            rc_date     = rc_info[0]
            rc_name_str = rc_info[1]
            mes_cobro   = rc_date[:7] if rc_date else inv_month

            if mes_cobro:
                all_months.add(mes_cobro)

            detail_rows.append({
                "mes_cobro":     mes_cobro,
                "fecha_factura": inv_date,
                "factura":       inv_name,
                "moneda":        moneda,
                "fecha_pago":    rc_date,
                "recibo":        rc_name_str,
                "cliente":       cname,
                "vendedor":      seller_name,
                "vendedor_id":   seller_id,
                "total_factura": money_to_float(inv_total),
                "nc_ids":        nc_ids,
                "total_nc":      money_to_float(nc_total),
                "saldo_neto":    money_to_float(saldo_neto),
                "comision":      money_to_float(commission),
            })

            sm = seller_month[seller_id][mes_cobro]
            sm["saldo_neto"]       += saldo_neto
            sm["comision_pagada"]  += commission
            sm["facturas_pagadas"] += 1

            st["facturas_pagadas"]   += 1
            st["comision_pagada"]    += commission
        else:
            # Factura sin recibo de caja — comisión pendiente
            dias_sin_pago = (date.today() - date.fromisoformat(inv_date)).days if inv_date else 0
            pending_rows.append({
                "fecha_factura":    inv_date,
                "factura":          inv_name,
                "moneda":           moneda,
                "cliente":          cname,
                "vendedor":         seller_name,
                "total_factura":    money_to_float(inv_total),
                "nc_ids":           nc_ids,
                "total_nc":         money_to_float(nc_total),
                "saldo_neto":       money_to_float(saldo_neto),
                "comision_pte":     money_to_float(commission),
                "dias_sin_pago":    dias_sin_pago,
                "balance_api":      money_to_float(balance_api),
            })

            st["facturas_pendientes"] += 1
            st["comision_pendiente"]  += commission

    detail_rows.sort(key=lambda r: (r["mes_cobro"], r["vendedor"], r["fecha_factura"]))
    pending_rows.sort(key=lambda r: r["fecha_factura"])

    sorted_months = sorted(all_months)

    summary_rows = [
        {
            "vendedor":            v["vendedor"],
            "facturas_total":      v["facturas_total"],
            "facturas_pagadas":    v["facturas_pagadas"],
            "facturas_pendientes": v["facturas_pendientes"],
            "total_factura":       money_to_float(v["total_factura"]),
            "total_nc":            money_to_float(v["total_nc"]),
            "saldo_neto":          money_to_float(v["saldo_neto"]),
            "comision_pagada":     money_to_float(v["comision_pagada"]),
            "comision_pendiente":  money_to_float(v["comision_pendiente"]),
            "comision_total":      money_to_float(v["comision_pagada"] + v["comision_pendiente"]),
        }
        for v in sorted(seller_total.values(), key=lambda x: -(x["comision_pagada"] + x["comision_pendiente"]))
    ]

    monthly_rows: list[dict[str, Any]] = []
    for sid, months in seller_month.items():
        sname = seller_total[sid]["vendedor"]
        for m in sorted_months:
            data = months.get(m)
            if not data or data["saldo_neto"] <= 0:
                continue
            monthly_rows.append({
                "mes":            m,
                "vendedor":       sname,
                "facturas":       data["facturas_pagadas"],
                "saldo_neto":     money_to_float(data["saldo_neto"]),
                "comision":       money_to_float(data["comision_pagada"]),
            })
    monthly_rows.sort(key=lambda r: (r["mes"], r["vendedor"]))

    # ── Notas crédito sin factura asociada ────────────────────────────────────
    orphan_rows: list[dict[str, Any]] = []
    for nc in orphan_ncs:
        nc_date_str = safe_text(nc.get("date"))
        nc_date = nc_date_str[:10] if nc_date_str else ""
        c    = nc.get("customer") if isinstance(nc.get("customer"), dict) else {}
        cid  = safe_text(c.get("id"))
        cname = customer_display_name(
            customers_map.get(cid),
            safe_text(c.get("identification")) or cid or "Sin cliente",
        )
        inv_ref      = nc.get("invoice")
        inv_ref_name = safe_text(inv_ref.get("name")) if isinstance(inv_ref, dict) else ""
        orphan_rows.append({
            "fecha_nc":     nc_date,
            "nota_credito": safe_text(nc.get("name")),
            "cliente":      cname,
            "factura_ref":  inv_ref_name,
            "total":        money_to_float(to_decimal(nc.get("total"))),
        })
    orphan_rows.sort(key=lambda r: r["fecha_nc"])

    return {
        "meta": {
            "from_date":         from_date.isoformat(),
            "to_date":           to_date.isoformat(),
            "commission_rate":   "3%",
            "generated_at":      datetime.now(timezone.utc).isoformat(),
            "invoices":              len(inv_map),
            "invoices_pagadas":      len(detail_rows),
            "invoices_pendientes":   len(pending_rows),
            "credit_notes":          len(raw_ncs),
            "orphan_ncs":            len(orphan_ncs),
            "sellers":               len(seller_total),
            "comision_pagada_total": money_to_float(sum(v["comision_pagada"]    for v in seller_total.values())),
            "comision_pte_total":    money_to_float(sum(v["comision_pendiente"] for v in seller_total.values())),
        },
        "summary":      summary_rows,
        "by_month":     monthly_rows,
        "detail":       detail_rows,
        "pending":      pending_rows,
        "orphan_ncs":   orphan_rows,
        "seller_month": seller_month,
        "seller_total": seller_total,
        "months":       sorted_months,
    }


# ── Hojas Excel ────────────────────────────────────────────────────────────────

def _sheet_resumen(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Resumen")
    ws.freeze_panes = "A3"
    meta = report["meta"]

    ncols = 10
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value = (
        f"Comisiones por vendedor  |  "
        f"{meta['from_date']} → {meta['to_date']}  |  "
        f"Tasa: {meta['commission_rate']}  |  "
        f"Base: saldo neto de TODAS las facturas del período"
    )
    t.font = Font(bold=True, size=13, color=WHITE)
    t.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    cols = [
        "Vendedor",
        "Fact. Total", "Fact. Pagadas", "Fact. Pendientes",
        "Total Factura", "Total NC", "Saldo Neto",
        "Comisión Pagada", "Comisión Pendiente", "Comisión Total (3%)",
    ]
    for ci, c in enumerate(cols, 1):
        _hcell(ws.cell(2, ci), c, COLOR_SUB)
    ws.row_dimensions[2].height = 22

    totals = {k: 0 for k in ["facturas_total", "facturas_pagadas", "facturas_pendientes"]}
    totals.update({k: 0.0 for k in ["total_factura", "total_nc", "saldo_neto",
                                     "comision_pagada", "comision_pendiente", "comision_total"]})
    for ri, row in enumerate(report["summary"], 3):
        _dcell(ws.cell(ri, 1), row["vendedor"])
        _dcell(ws.cell(ri, 2), row["facturas_total"],      "center")
        _dcell(ws.cell(ri, 3), row["facturas_pagadas"],    "center")
        _dcell(ws.cell(ri, 4), row["facturas_pendientes"], "center")
        _mcell(ws.cell(ri, 5), row["total_factura"])
        _mcell(ws.cell(ri, 6), row["total_nc"])
        ws.cell(ri, 6).font = Font(color="C00000")
        _mcell(ws.cell(ri, 7), row["saldo_neto"])
        _mcell(ws.cell(ri, 8), row["comision_pagada"])
        ws.cell(ri, 8).font = Font(bold=True, color="1F4E78")
        _mcell(ws.cell(ri, 9), row["comision_pendiente"])
        ws.cell(ri, 9).font = Font(color="E26B0A")
        _mcell(ws.cell(ri, 10), row["comision_total"])
        ws.cell(ri, 10).font = Font(bold=True, color="1F4E78")
        for k in ["facturas_total", "facturas_pagadas", "facturas_pendientes"]:
            totals[k] += row[k]
        for k in ["total_factura", "total_nc", "saldo_neto",
                  "comision_pagada", "comision_pendiente", "comision_total"]:
            totals[k] += row[k]

    tr = len(report["summary"]) + 3
    _tcell(ws.cell(tr, 1), "TOTAL")
    _tcell(ws.cell(tr, 2), totals["facturas_total"])
    _tcell(ws.cell(tr, 3), totals["facturas_pagadas"])
    _tcell(ws.cell(tr, 4), totals["facturas_pendientes"])
    _tcell(ws.cell(tr, 5), round(totals["total_factura"],       2), is_money=True)
    _tcell(ws.cell(tr, 6), round(totals["total_nc"],            2), is_money=True)
    _tcell(ws.cell(tr, 7), round(totals["saldo_neto"],          2), is_money=True)
    _tcell(ws.cell(tr, 8), round(totals["comision_pagada"],     2), is_money=True)
    _tcell(ws.cell(tr, 9), round(totals["comision_pendiente"],  2), is_money=True)
    _tcell(ws.cell(tr, 10), round(totals["comision_total"],     2), is_money=True)
    ws.cell(tr, 10).font = Font(bold=True, color="1F4E78")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 22


def _sheet_pivot(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Pivot Comisiones")
    ws.freeze_panes = "B2"

    months  = report["months"]
    sm      = report["seller_month"]
    st      = report["seller_total"]
    sellers = sorted(st.values(), key=lambda x: -(x["comision_pagada"] + x["comision_pendiente"]))

    if not months or not sellers:
        ws["A1"] = "Sin datos"
        return

    _hcell(ws.cell(1, 1), "Vendedor")
    for ci, m in enumerate(months, 2):
        _hcell(ws.cell(1, ci), m, COLOR_SUB)
    total_col = len(months) + 2
    _hcell(ws.cell(1, total_col), "TOTAL")

    col_totals: dict[int, float] = defaultdict(float)
    grand = 0.0

    for ri, seller in enumerate(sellers, 2):
        sid = seller["vendedor_id"]
        _dcell(ws.cell(ri, 1), seller["vendedor"])
        row_total = 0.0
        for ci, m in enumerate(months, 2):
            val = money_to_float(sm[sid][m]["comision_pagada"]) if m in sm[sid] else 0.0
            if val:
                _mcell(ws.cell(ri, ci), val)
            else:
                c = ws.cell(ri, ci)
                c.value = "—"
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = _border()
            row_total += val
            col_totals[ci] += val
        _tcell(ws.cell(ri, total_col), round(row_total, 2), is_money=True)
        ws.cell(ri, total_col).font = Font(bold=True, color="1F4E78")
        grand += row_total

    tr = len(sellers) + 2
    _tcell(ws.cell(tr, 1), "TOTAL")
    for ci in range(2, len(months) + 2):
        _tcell(ws.cell(tr, ci), round(col_totals[ci], 2), is_money=True)
    _tcell(ws.cell(tr, total_col), round(grand, 2), is_money=True)
    ws.cell(tr, total_col).font = Font(bold=True, color="1F4E78")

    ws.column_dimensions["A"].width = 32
    for ci in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18


def _sheet_por_mes(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Por Mes")
    ws.freeze_panes = "A2"

    cols = ["Mes Cobro", "Vendedor", "N° Facturas", "Saldo Neto", "Comisión (3%)"]
    for ci, c in enumerate(cols, 1):
        _hcell(ws.cell(1, ci), c)

    for ri, row in enumerate(report["by_month"], 2):
        _dcell(ws.cell(ri, 1), row["mes"],      "center")
        _dcell(ws.cell(ri, 2), row["vendedor"])
        _dcell(ws.cell(ri, 3), row["facturas"], "center")
        _mcell(ws.cell(ri, 4), row["saldo_neto"])
        _mcell(ws.cell(ri, 5), row["comision"])
        ws.cell(ri, 5).font = Font(bold=True, color="1F4E78")

    _add_table(ws, len(report["by_month"]), len(cols), "ComisionesPorMes")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22


def _sheet_detalle(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Detalle Facturas")
    ws.freeze_panes = "A2"

    cols = [
        "Mes Cobro", "Fecha Factura", "Factura", "Moneda",
        "Fecha Pago", "Recibo RC",
        "Cliente", "Vendedor",
        "Total Factura (COP)", "ID Nota Crédito", "Total NC (COP)",
        "Saldo Final (COP)", "Comisión (3%)",
    ]
    for ci, c in enumerate(cols, 1):
        _hcell(ws.cell(1, ci), c)

    for ri, row in enumerate(report["detail"], 2):
        _dcell(ws.cell(ri,  1), row["mes_cobro"],     "center")
        _dcell(ws.cell(ri,  2), row["fecha_factura"], "center")
        _dcell(ws.cell(ri,  3), row["factura"])
        _dcell(ws.cell(ri,  4), row.get("moneda", "COP"), "center")
        _dcell(ws.cell(ri,  5), row["fecha_pago"],    "center")
        _dcell(ws.cell(ri,  6), row["recibo"])
        _dcell(ws.cell(ri,  7), row["cliente"])
        _dcell(ws.cell(ri,  8), row["vendedor"])
        _mcell(ws.cell(ri,  9), row["total_factura"])
        _dcell(ws.cell(ri, 10), row["nc_ids"] or "—")
        if row["total_nc"]:
            _mcell(ws.cell(ri, 11), row["total_nc"])
            ws.cell(ri, 11).font = Font(color="C00000")
        else:
            c = ws.cell(ri, 11)
            c.value = "—"
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()
        _mcell(ws.cell(ri, 12), row["saldo_neto"])
        ws.cell(ri, 12).font = Font(bold=True)
        if row["comision"] > 0:
            _mcell(ws.cell(ri, 13), row["comision"])
            ws.cell(ri, 13).font = Font(bold=True, color="1F4E78")
        else:
            c = ws.cell(ri, 13)
            c.value = "—"
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()

    _add_table(ws, len(report["detail"]), len(cols), "DetalleFacturas")
    widths = [10, 14, 14, 8, 12, 14, 30, 28, 20, 22, 18, 20, 20]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _sheet_pendientes(wb: Workbook, report: dict[str, Any]) -> None:
    rows = report["pending"]
    ws = wb.create_sheet("Facturas Pendientes")
    ws.freeze_panes = "A2"

    cols = [
        "Fecha Factura", "Factura", "Moneda", "Cliente", "Vendedor",
        "Total Factura (COP)", "ID Nota Crédito", "Total NC (COP)",
        "Saldo Neto (COP)", "Comisión Pendiente", "Días Sin Pago",
    ]
    for ci, c in enumerate(cols, 1):
        _hcell(ws.cell(1, ci), c, COLOR_HEADER)

    if not rows:
        ws.merge_cells(f"A2:{get_column_letter(len(cols))}2")
        c = ws["A2"]
        c.value = "Todas las facturas del período han sido cobradas"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    else:
        for ri, row in enumerate(rows, 2):
            _dcell(ws.cell(ri,  1), row["fecha_factura"], "center")
            _dcell(ws.cell(ri,  2), row["factura"])
            _dcell(ws.cell(ri,  3), row.get("moneda", "COP"), "center")
            _dcell(ws.cell(ri,  4), row["cliente"])
            _dcell(ws.cell(ri,  5), row["vendedor"])
            _mcell(ws.cell(ri,  6), row["total_factura"])
            _dcell(ws.cell(ri,  7), row["nc_ids"] or "—")
            if row["total_nc"]:
                _mcell(ws.cell(ri, 8), row["total_nc"])
                ws.cell(ri, 8).font = Font(color="C00000")
            else:
                c = ws.cell(ri, 8)
                c.value = "—"
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = _border()
            _mcell(ws.cell(ri,  9), row["saldo_neto"])
            ws.cell(ri, 9).font = Font(bold=True, color="C00000")
            _mcell(ws.cell(ri, 10), row["comision_pte"])
            ws.cell(ri, 10).font = Font(color="E26B0A")
            dias = row["dias_sin_pago"]
            _dcell(ws.cell(ri, 11), dias, "center")
            if dias > 90:
                ws.cell(ri, 10).font = Font(bold=True, color="C00000")
            elif dias > 30:
                ws.cell(ri, 10).font = Font(color="E26B0A")
        _add_table(ws, len(rows), len(cols), "FacturasPendientes")

    widths = [14, 14, 8, 30, 28, 20, 22, 18, 20, 20, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _sheet_nc_sin_factura(wb: Workbook, report: dict[str, Any]) -> None:
    rows = report["orphan_ncs"]
    ws = wb.create_sheet("NC Sin Factura")
    ws.freeze_panes = "A2"

    cols = ["Fecha NC", "Nota Crédito", "Cliente", "Factura Referenciada", "Total NC"]
    for ci, c in enumerate(cols, 1):
        _hcell(ws.cell(1, ci), c)

    if not rows:
        ws.merge_cells("A2:E2")
        c = ws["A2"]
        c.value = "No hay notas crédito sin factura asociada en el período"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    else:
        for ri, row in enumerate(rows, 2):
            _dcell(ws.cell(ri, 1), row["fecha_nc"],     "center")
            _dcell(ws.cell(ri, 2), row["nota_credito"])
            _dcell(ws.cell(ri, 3), row["cliente"])
            _dcell(ws.cell(ri, 4), row["factura_ref"] or "Sin referencia")
            _mcell(ws.cell(ri, 5), row["total"])
            ws.cell(ri, 5).font = Font(color="C00000")
        _add_table(ws, len(rows), len(cols), "NCSinFactura")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20


def write_commission_workbook(output_path: Path, report: dict[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_resumen(wb, report)
    _sheet_pivot(wb, report)
    _sheet_por_mes(wb, report)
    _sheet_detalle(wb, report)
    _sheet_pendientes(wb, report)
    _sheet_nc_sin_factura(wb, report)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── CLI ────────────────────────────────────────────────────────────────────────

def main() -> int:
    p = argparse.ArgumentParser(
        description="Comisiones de vendedores — 3% sobre saldo neto de facturas."
    )
    p.add_argument("--from-date", required=True, help="Fecha inicial YYYY-MM-DD")
    p.add_argument("--to-date",   required=True, help="Fecha final YYYY-MM-DD")
    p.add_argument("--output",    required=True, help="Ruta del .xlsx de salida")
    p.add_argument("--timeout",   type=int, default=60)
    args = p.parse_args()

    try:
        from_date = parse_iso_date(args.from_date, "--from-date")
        to_date   = parse_iso_date(args.to_date,   "--to-date")
        if from_date > to_date:
            raise SiigoApiError("--from-date no puede ser mayor que --to-date")

        config = build_siigo_config()
        print(
            f"\nGenerando comisiones {from_date} → {to_date} "
            f"(base: saldo neto | timing: recibo de caja)\n"
        )

        report = build_commission_report(config, from_date, to_date, timeout=args.timeout)

        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        write_commission_workbook(output_path, report)

        meta    = report["meta"]
        com_pag = meta["comision_pagada_total"]
        com_pte = meta["comision_pte_total"]

        print(f"\n{'─'*58}")
        print(f"  Período              : {meta['from_date']} → {meta['to_date']}")
        print(f"  Facturas             : {meta['invoices']} total  |  {meta['invoices_pagadas']} pagadas  |  {meta['invoices_pendientes']} pendientes")
        print(f"  Notas crédito        : {meta['credit_notes']} ({meta['orphan_ncs']} sin factura)")
        print(f"  Vendedores           : {meta['sellers']}")
        print(f"  Tasa                 : {meta['commission_rate']}")
        print(f"  Comisión pagada      : $ {com_pag:,.2f}")
        print(f"  Comisión pendiente   : $ {com_pte:,.2f}")
        print(f"  Comisión total       : $ {com_pag + com_pte:,.2f}")
        print(f"{'─'*58}")
        print()
        for row in report["summary"]:
            print(f"  {row['vendedor']:<34}  pagada $ {row['comision_pagada']:>14,.2f}  |  pendiente $ {row['comision_pendiente']:>14,.2f}  |  total $ {row['comision_total']:>14,.2f}")
        if meta["orphan_ncs"]:
            print(f"\n  ⚠  {meta['orphan_ncs']} notas crédito sin factura asociada (ver hoja NC Sin Factura)")
        print(f"\n  Archivo       : {output_path}\n")
        return 0

    except SiigoApiError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
