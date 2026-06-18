#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import calendar
from datetime import date
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
import os
import re
import ssl
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib import error, parse, request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from siigo_core import fetch_credit_notes, invoice_exchange_rate


DEFAULT_BASE_URL = "https://api.siigo.com"
AUTH_PATH_CANDIDATES = ("/auth", "/v1/auth")
SYSTEM_CA_CANDIDATES = (
    "/etc/ssl/cert.pem",
    "/private/etc/ssl/cert.pem",
)
MONEY_PLACES = Decimal("0.01")


class SiigoApiError(RuntimeError):
    pass


@dataclass
class Config:
    username: str
    access_key: str
    partner_id: str
    base_url: str = DEFAULT_BASE_URL
    auth_header: str | None = None


def load_dotenv(env_path: Path) -> None:
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'").strip('"')
        os.environ.setdefault(key, value)


def require_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise SiigoApiError(f"Falta la variable de entorno obligatoria: {name}")
    return value


def build_config() -> Config:
    load_dotenv(Path(".env"))
    return Config(
        username=require_env("SIIGO_USERNAME"),
        access_key=require_env("SIIGO_ACCESS_KEY"),
        partner_id=require_env("SIIGO_PARTNER_ID"),
        base_url=os.getenv("SIIGO_BASE_URL", DEFAULT_BASE_URL).rstrip("/"),
        auth_header=os.getenv("SIIGO_AUTHORIZATION_HEADER", "").strip() or None,
    )


def json_request(
    method: str,
    url: str,
    headers: dict[str, str],
    payload: dict[str, Any] | None = None,
    timeout: int = 30,
) -> tuple[int, dict[str, Any] | list[Any] | str | None]:
    data = None
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")

    req = request.Request(url=url, data=data, method=method.upper())
    for key, value in headers.items():
        req.add_header(key, value)

    ssl_context = build_ssl_context()

    try:
        with request.urlopen(req, timeout=timeout, context=ssl_context) as response:
            status = response.status
            raw_body = response.read()
            return status, decode_body(raw_body, response.headers.get("Content-Type", ""))
    except error.HTTPError as exc:
        raw_body = exc.read()
        body = decode_body(raw_body, exc.headers.get("Content-Type", ""))
        raise SiigoApiError(format_http_error(exc.code, url, body)) from exc
    except error.URLError as exc:
        raise SiigoApiError(f"No fue posible conectarse a Siigo: {exc.reason}") from exc


def decode_body(raw_body: bytes, content_type: str) -> dict[str, Any] | list[Any] | str | None:
    if not raw_body:
        return None

    text = raw_body.decode("utf-8", errors="replace")
    if "application/json" in content_type.lower():
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text
    return text


def build_ssl_context() -> ssl.SSLContext:
    custom_cafile = os.getenv("SIIGO_CA_FILE", "").strip()
    if custom_cafile:
        cafile = Path(custom_cafile)
        if not cafile.exists():
            raise SiigoApiError(f"El archivo SIIGO_CA_FILE no existe: {cafile}")
        return ssl.create_default_context(cafile=str(cafile))

    default_paths = ssl.get_default_verify_paths()
    if default_paths.cafile and Path(default_paths.cafile).exists():
        return ssl.create_default_context()

    for candidate in SYSTEM_CA_CANDIDATES:
        if Path(candidate).exists():
            return ssl.create_default_context(cafile=candidate)

    return ssl.create_default_context()


def format_http_error(status: int, url: str, body: Any) -> str:
    if isinstance(body, (dict, list)):
        serialized = json.dumps(body, ensure_ascii=False)
    elif body is None:
        serialized = "(sin cuerpo)"
    else:
        serialized = str(body)
    return f"Error HTTP {status} al consumir {url}: {serialized}"


def obtain_token(config: Config, timeout: int = 30) -> tuple[str, str]:
    payload = {
        "username": config.username,
        "access_key": config.access_key,
    }

    last_error: SiigoApiError | None = None
    for auth_path in AUTH_PATH_CANDIDATES:
        url = f"{config.base_url}{auth_path}"
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Partner-Id": config.partner_id,
        }
        if config.auth_header:
            headers["Authorization"] = config.auth_header

        try:
            _, body = json_request("POST", url, headers, payload=payload, timeout=timeout)
        except SiigoApiError as exc:
            last_error = exc
            continue

        if not isinstance(body, dict) or "access_token" not in body:
            raise SiigoApiError(
                f"La respuesta de autenticación no tiene el formato esperado: {body!r}"
            )

        token = str(body["access_token"]).strip()
        token_type = str(body.get("token_type", "Bearer")).strip() or "Bearer"
        if not token:
            raise SiigoApiError("Siigo respondió sin access_token.")
        return token, token_type

    if last_error is not None:
        raise last_error
    raise SiigoApiError("No fue posible obtener el token de autenticación.")


def normalize_path(path: str) -> str:
    cleaned = path.strip().lstrip("/")
    if "://" in cleaned:
        raise SiigoApiError("El path debe ser relativo, por ejemplo: v1/customers")
    if cleaned in {"auth", "v1/auth"}:
        raise SiigoApiError("La autenticación ya la maneja el script internamente.")
    if not cleaned.startswith("v1/"):
        raise SiigoApiError("Por seguridad el path debe iniciar con 'v1/'.")
    return cleaned


def parse_params(values: list[str]) -> dict[str, str]:
    params: dict[str, str] = {}
    for value in values:
        if "=" not in value:
            raise SiigoApiError(
                f"Parámetro inválido: {value}. Usa el formato --param clave=valor"
            )
        key, raw = value.split("=", 1)
        key = key.strip()
        if not key:
            raise SiigoApiError(f"Parámetro inválido: {value}")
        params[key] = raw.strip()
    return params


def authorization_candidates(token: str, token_type: str) -> list[str]:
    candidates: list[str] = []
    typed = token if " " in token else f"{token_type} {token}".strip()
    raw = token

    for candidate in (typed, raw):
        if candidate and candidate not in candidates:
            candidates.append(candidate)
    return candidates


def build_get_url(config: Config, path: str, params: dict[str, str]) -> str:
    query = parse.urlencode(params)
    url = f"{config.base_url}/{normalize_path(path)}"
    if query:
        url = f"{url}?{query}"
    return url


def get_read_headers(config: Config, authorization: str) -> dict[str, str]:
    return {
        "Accept": "application/json",
        "Authorization": authorization,
        "Partner-Id": config.partner_id,
    }


def readonly_get_url(
    config: Config,
    url: str,
    timeout: int = 30,
    token: str | None = None,
    token_type: str | None = None,
) -> tuple[int, dict[str, Any] | list[Any] | str | None]:
    if token is None or token_type is None:
        token, token_type = obtain_token(config, timeout=timeout)

    last_error: SiigoApiError | None = None
    for authorization in authorization_candidates(token, token_type):
        headers = get_read_headers(config, authorization)
        try:
            return json_request("GET", url, headers, timeout=timeout)
        except SiigoApiError as exc:
            last_error = exc
            continue

    if last_error is not None:
        raise last_error
    raise SiigoApiError("La consulta GET no se pudo completar.")


def readonly_get(
    config: Config,
    path: str,
    params: dict[str, str],
    timeout: int = 30,
) -> tuple[int, dict[str, Any] | list[Any] | str | None]:
    token, token_type = obtain_token(config, timeout=timeout)
    url = build_get_url(config, path, params)
    return readonly_get_url(
        config,
        url,
        timeout=timeout,
        token=token,
        token_type=token_type,
    )


def fetch_paginated_results(
    config: Config,
    path: str,
    params: dict[str, str],
    timeout: int = 30,
    max_pages: int | None = None,
) -> dict[str, Any]:
    normalized_path = normalize_path(path)
    current_url = build_get_url(config, normalized_path, params)
    token, token_type = obtain_token(config, timeout=timeout)

    seen_urls: set[str] = set()
    items: list[Any] = []
    pages_fetched = 0
    reported_total_results: int | None = None

    while current_url:
        if current_url in seen_urls:
            raise SiigoApiError(f"Se detectó un ciclo de paginación en {current_url}")
        seen_urls.add(current_url)

        status, body = readonly_get_url(
            config,
            current_url,
            timeout=timeout,
            token=token,
            token_type=token_type,
        )
        if status != 200:
            raise SiigoApiError(f"Respuesta inesperada HTTP {status} en {current_url}")
        if not isinstance(body, dict):
            raise SiigoApiError("La exportación esperaba una respuesta JSON tipo objeto.")

        results = body.get("results")
        if not isinstance(results, list):
            raise SiigoApiError(
                "La exportación paginada solo funciona con endpoints que respondan con 'results'."
            )

        items.extend(results)
        pages_fetched += 1

        pagination = body.get("pagination")
        if (
            reported_total_results is None
            and isinstance(pagination, dict)
            and isinstance(pagination.get("total_results"), int)
        ):
            reported_total_results = pagination["total_results"]

        if max_pages is not None and pages_fetched >= max_pages:
            break

        current_url = extract_next_href(body) or build_next_url_from_pagination(current_url, body)

    return {
        "path": normalized_path,
        "params": params,
        "pages_fetched": pages_fetched,
        "reported_total_results": reported_total_results,
        "results": items,
    }


def extract_next_href(body: dict[str, Any]) -> str | None:
    for links_key in ("_links", "__links"):
        links = body.get(links_key)
        if not isinstance(links, dict):
            continue
        next_link = links.get("next")
        if not isinstance(next_link, dict):
            continue
        href = next_link.get("href")
        if isinstance(href, str) and href.strip():
            return href.strip()
    return None


def build_next_url_from_pagination(current_url: str, body: dict[str, Any]) -> str | None:
    pagination = body.get("pagination")
    results = body.get("results")
    if not isinstance(pagination, dict) or not isinstance(results, list):
        return None

    page = pagination.get("page")
    page_size = pagination.get("page_size")
    total_results = pagination.get("total_results")
    if not all(isinstance(value, int) for value in (page, page_size, total_results)):
        return None
    if page < 1 or page_size < 1 or total_results <= page * page_size:
        return None

    parsed = parse.urlsplit(current_url)
    query = dict(parse.parse_qsl(parsed.query, keep_blank_values=True))
    query["page"] = str(page + 1)
    query.setdefault("page_size", str(page_size))
    new_query = parse.urlencode(query)
    return parse.urlunsplit((parsed.scheme, parsed.netloc, parsed.path, new_query, parsed.fragment))


def export_paginated_get(
    config: Config,
    path: str,
    params: dict[str, str],
    timeout: int = 30,
    max_pages: int | None = None,
) -> dict[str, Any]:
    fetched = fetch_paginated_results(
        config,
        path,
        params,
        timeout=timeout,
        max_pages=max_pages,
    )
    return {
        "summary": {
            "path": fetched["path"],
            "params": params,
            "pages_fetched": fetched["pages_fetched"],
            "exported_results": len(fetched["results"]),
            "reported_total_results": fetched["reported_total_results"],
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        },
        "results": fetched["results"],
    }


def parse_iso_date(value: str, label: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise SiigoApiError(f"{label} debe tener formato YYYY-MM-DD") from exc


def shift_months(base_date: date, months: int) -> date:
    year = base_date.year
    month = base_date.month - months
    while month <= 0:
        month += 12
        year -= 1
    day = min(base_date.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def resolve_report_range(
    months: int,
    from_date_raw: str | None,
    to_date_raw: str | None,
) -> tuple[date, date]:
    end_date = parse_iso_date(to_date_raw, "--to-date") if to_date_raw else date.today()
    if from_date_raw:
        start_date = parse_iso_date(from_date_raw, "--from-date")
    else:
        start_date = shift_months(end_date, months)

    if start_date > end_date:
        raise SiigoApiError("La fecha inicial no puede ser mayor a la fecha final.")
    return start_date, end_date


def date_bounds_to_rfc3339(start_date: date, end_date: date) -> tuple[str, str]:
    return (
        f"{start_date.isoformat()}T00:00:00Z",
        f"{end_date.isoformat()}T23:59:59Z",
    )


def to_decimal(value: Any) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str) and value.strip():
        try:
            return Decimal(value.strip())
        except InvalidOperation:
            return Decimal("0")
    return Decimal("0")


def money_to_float(value: Decimal) -> float:
    return float(value.quantize(MONEY_PLACES, rounding=ROUND_HALF_UP))


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def dedupe_repeated_phrase(text: str) -> str:
    words = [word for word in text.split() if word]
    if len(words) % 2 == 0 and words:
        midpoint = len(words) // 2
        if [word.casefold() for word in words[:midpoint]] == [word.casefold() for word in words[midpoint:]]:
            return " ".join(words[:midpoint])
    return text


_PLACEHOLDER_NAMES: frozenset[str] = frozenset({
    "noaplica", "n/a", "na", "sinnombre", "ninguno", "sinregistro",
})


def _is_placeholder(text: str) -> bool:
    return text.lower().replace(" ", "") in _PLACEHOLDER_NAMES


def customer_display_name(customer_body: dict[str, Any] | None, fallback_name: str) -> str:
    if isinstance(customer_body, dict):
        raw_name = customer_body.get("name")
        if isinstance(raw_name, list):
            joined = " ".join(part.strip() for part in raw_name if isinstance(part, str) and part.strip())
            if joined and not _is_placeholder(joined):
                return joined
        if isinstance(raw_name, str) and raw_name.strip() and not _is_placeholder(raw_name.strip()):
            return raw_name.strip()

        commercial_name = safe_text(customer_body.get("commercial_name"))
        if commercial_name and not _is_placeholder(commercial_name):
            return commercial_name

    return fallback_name


def user_display_name(user_body: dict[str, Any] | None, fallback_name: str) -> str:
    if isinstance(user_body, dict):
        first_name = safe_text(user_body.get("first_name"))
        last_name = safe_text(user_body.get("last_name"))
        if first_name and last_name and first_name.casefold() == last_name.casefold():
            return first_name
        full_name = dedupe_repeated_phrase(" ".join(part for part in (first_name, last_name) if part).strip())
        if full_name:
            return full_name

        for key in ("username", "email", "identification"):
            value = safe_text(user_body.get(key))
            if value:
                return value

    return fallback_name


def fetch_customer_details(
    config: Config,
    customer_ids: dict[str, dict[str, str]],
    timeout: int = 30,
) -> tuple[dict[str, dict[str, Any]], list[dict[str, str]]]:
    if not customer_ids:
        return {}, []

    token, token_type = obtain_token(config, timeout=timeout)
    resolved: dict[str, dict[str, Any]] = {}
    failures: list[dict[str, str]] = []

    for customer_id, hints in sorted(customer_ids.items()):
        url = f"{config.base_url}/{normalize_path(f'v1/customers/{customer_id}')}"
        try:
            _, body = readonly_get_url(
                config,
                url,
                timeout=timeout,
                token=token,
                token_type=token_type,
            )
            if isinstance(body, dict):
                resolved[customer_id] = body
                continue
            raise SiigoApiError("Respuesta inesperada al consultar cliente.")
        except SiigoApiError as exc:
            failures.append(
                {
                    "customer_id": customer_id,
                    "customer_identification": hints.get("identification", ""),
                    "error": str(exc),
                }
            )

    return resolved, failures


def fetch_users_map(
    config: Config,
    timeout: int = 30,
) -> tuple[dict[str, dict[str, Any]], dict[str, int]]:
    fetched = fetch_paginated_results(
        config,
        "v1/users",
        {
            "page": "1",
            "page_size": "100",
        },
        timeout=timeout,
    )

    users_map: dict[str, dict[str, Any]] = {}
    for user in fetched["results"]:
        if not isinstance(user, dict):
            continue
        user_id = safe_text(user.get("id"))
        if user_id:
            users_map[user_id] = user

    return users_map, {
        "pages_fetched": fetched["pages_fetched"],
        "reported_total_results": fetched["reported_total_results"] or len(users_map),
    }


def write_csv(output_path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Use UTF-8 with BOM plus locale-friendly CSV formatting for Spanish Excel.
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: format_csv_value(row.get(key)) for key in fieldnames})


def format_csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        decimal_value = Decimal(str(value)).quantize(MONEY_PLACES, rounding=ROUND_HALF_UP)
        return f"{decimal_value:.2f}".replace(".", ",")
    if isinstance(value, Decimal):
        decimal_value = value.quantize(MONEY_PLACES, rounding=ROUND_HALF_UP)
        return f"{decimal_value:.2f}".replace(".", ",")
    return str(value)


def excel_cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value


def style_excel_sheet(
    worksheet: Any,
    rows: list[dict[str, Any]],
    columns: list[str],
    table_name: str,
    numeric_columns: set[str] | None = None,
    date_columns: set[str] | None = None,
    headers: list[str] | None = None,
) -> None:
    numeric_columns = numeric_columns or set()
    date_columns = date_columns or set()
    display_headers = headers if headers is not None else columns

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index in range(2, len(rows) + 2):
        for col_index, column_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            if column_name in numeric_columns:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif column_name in date_columns:
                cell.number_format = 'yyyy-mm-dd'
            elif isinstance(cell.value, bool):
                cell.alignment = Alignment(horizontal="center")

    for col_index, (column_name, header_label) in enumerate(zip(columns, display_headers), start=1):
        max_len = len(header_label)
        for row in rows:
            raw = row.get(column_name, "")
            display = "" if raw is None else str(raw)
            max_len = max(max_len, len(display))
        worksheet.column_dimensions[get_column_letter(col_index)].width = min(max_len + 2, 40)

    if rows:
        end_cell = f"{get_column_letter(len(columns))}{len(rows) + 1}"
        table = Table(displayName=table_name, ref=f"A1:{end_cell}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)


_COL_ES: dict[str, str] = {
    "metric":                 "Métrica",
    "value":                  "Valor",
    "date":                   "Fecha",
    "month":                  "Mes",
    "invoice_name":           "Factura",
    "nc_name":                "Nota Crédito",
    "nc_id":                  "ID Nota Crédito",
    "linked_invoice":         "Factura Referenciada",
    "customer_id":            "ID Cliente",
    "customer_identification":"NIT / CC",
    "customer_branch_office": "Sucursal",
    "customer_name":          "Cliente",
    "seller_id":              "ID Vendedor",
    "seller_name":            "Vendedor",
    "invoice_count":          "N° Facturas",
    "seller_count":           "N° Vendedores",
    "customer_count":         "N° Clientes",
    "total_sales":            "Total Ventas",
    "total":                  "Total",
    "balance":                "Saldo",
    "paid":                   "Pagado",
    "paid_amount":            "Pagado",
    "outstanding_balance":    "Saldo Pendiente",
    "credit_notes_total":     "Total Notas Crédito",
    "net_sales":              "Ventas Netas",
    "avg_invoice_total":      "Promedio Factura",
    "annulled":               "Anulada",
    "currency_code":          "Moneda",
    "bucket":                 "Rango",
    "total_balance":          "Saldo Total",
    "current_30":             "Corriente (0-30 d)",
    "overdue_31_60":          "Vencido 31-60 d",
    "overdue_61_90":          "Vencido 61-90 d",
    "overdue_91_plus":        "Vencido +90 d",
    "due_date":               "Fecha Vencimiento",
    "overdue_days":           "Días Vencidos",
}


def append_dict_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    table_name: str,
    numeric_columns: set[str] | None = None,
    date_columns: set[str] | None = None,
) -> None:
    worksheet = workbook.create_sheet(title=title)
    if not rows:
        worksheet["A1"] = "Sin datos"
        worksheet["A1"].font = Font(bold=True)
        return

    columns = list(rows[0].keys())
    headers = [_COL_ES.get(col, col) for col in columns]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([excel_cell_value(row.get(column)) for column in columns])

    style_excel_sheet(
        worksheet,
        rows,
        columns,
        table_name=table_name,
        numeric_columns=numeric_columns,
        date_columns=date_columns,
        headers=headers,
    )


_SUMMARY_ES: dict[str, str] = {
    "report_type":                   "Tipo de reporte",
    "from_date":                     "Fecha inicial",
    "to_date":                       "Fecha final",
    "date_start_rfc3339":            "Inicio (RFC3339)",
    "date_end_rfc3339":              "Fin (RFC3339)",
    "months_requested":              "Meses solicitados",
    "include_annulled":              "Incluir anuladas",
    "pages_fetched":                 "Páginas descargadas",
    "raw_invoices":                  "Facturas brutas",
    "report_invoices":               "Facturas en reporte",
    "annulled_invoices_in_range":    "Facturas anuladas en rango",
    "customers":                     "Clientes",
    "sellers":                       "Vendedores",
    "customer_seller_pairs":         "Pares cliente-vendedor",
    "total_sales":                   "Total ventas",
    "credit_notes_total":            "Total notas crédito",
    "net_sales":                     "Ventas netas",
    "outstanding_balance":           "Saldo pendiente",
    "reported_total_results":        "Total resultados reportados",
    "customer_lookup_failures":      "Clientes sin resolución",
    "users_pages_fetched":           "Páginas vendedores",
    "users_reported_total":          "Total vendedores reportados",
    "unresolved_sellers":            "Vendedores sin resolución",
    "credit_notes_count":            "Notas crédito",
    "generated_at_utc":              "Generado (UTC)",
    "report_invoices_with_balance":  "Facturas con saldo",
    "total_balance":                 "Saldo total cartera",
}


def build_summary_rows(summary: dict[str, Any]) -> list[dict[str, Any]]:
    return [{"metric": _SUMMARY_ES.get(key, key), "value": value} for key, value in summary.items()]


def average_amount(total: Decimal, count: int) -> float:
    if count <= 0:
        return 0.0
    return money_to_float(total / Decimal(count))


def _fetch_credit_notes_safe(
    config: Any,
    start_date: date,
    end_date: date,
    timeout: int = 30,
) -> list[dict[str, Any]]:
    try:
        return fetch_credit_notes(config, start_date, end_date, timeout=timeout)
    except SiigoApiError as exc:
        print(f"  Advertencia: no se pudieron obtener notas crédito: {exc}", file=sys.stderr)
        return []


def _nc_fx(nc: dict[str, Any], inv_fx_map: dict[str, Decimal] | None = None) -> Decimal:
    """Retorna el exchange rate COP de una NC usando su propia tasa de emisión."""
    return invoice_exchange_rate(nc)


def _build_nc_maps_by_customer(
    credit_notes: list[dict[str, Any]],
    inv_fx_map: dict[str, Decimal] | None = None,
) -> tuple[dict[str, Decimal], dict[str, Decimal]]:
    """Devuelve (nc_por_customer_key, nc_por_month|customer_key) en COP."""
    inv_fx_map = inv_fx_map or {}
    nc_customer: dict[str, Decimal] = {}
    nc_customer_month: dict[str, Decimal] = {}
    for nc in credit_notes:
        if nc.get("annulled"):
            continue
        nc_date = safe_text(nc.get("date"))
        if not nc_date:
            continue
        c = nc.get("customer") if isinstance(nc.get("customer"), dict) else {}
        cid = safe_text(c.get("id")) or "unknown"
        branch = safe_text(c.get("branch_office"))
        key = f"{cid}|{branch}"
        fx = _nc_fx(nc, inv_fx_map)
        total = to_decimal(nc.get("total")) * fx
        nc_customer[key] = nc_customer.get(key, Decimal("0")) + total
        month_key = f"{nc_date[:7]}|{key}"
        nc_customer_month[month_key] = nc_customer_month.get(month_key, Decimal("0")) + total
    return nc_customer, nc_customer_month


def _build_nc_detail_rows(
    credit_notes: list[dict[str, Any]],
    inv_fx_map: dict[str, Decimal] | None = None,
) -> list[dict[str, Any]]:
    inv_fx_map = inv_fx_map or {}
    rows: list[dict[str, Any]] = []
    for nc in credit_notes:
        if nc.get("annulled"):
            continue
        nc_date = safe_text(nc.get("date"))
        if not nc_date:
            continue
        c = nc.get("customer") if isinstance(nc.get("customer"), dict) else {}
        inv_ref = nc.get("invoice")
        fx = _nc_fx(nc, inv_fx_map)
        rows.append({
            "date": nc_date,
            "month": nc_date[:7],
            "nc_name": safe_text(nc.get("name")),
            "nc_id": safe_text(nc.get("id")),
            "customer_id": safe_text(c.get("id")),
            "customer_identification": safe_text(c.get("identification")),
            "customer_branch_office": safe_text(c.get("branch_office")),
            "linked_invoice": safe_text(inv_ref.get("name")) if isinstance(inv_ref, dict) else "",
            "total": money_to_float(to_decimal(nc.get("total")) * fx),
        })
    rows.sort(key=lambda r: (r["date"], r["nc_name"]))
    return rows


def build_sales_report(
    config: Config,
    months: int,
    from_date_raw: str | None,
    to_date_raw: str | None,
    include_annulled: bool,
    page_size: int,
    timeout: int = 30,
    max_pages: int | None = None,
) -> dict[str, Any]:
    start_date, end_date = resolve_report_range(months, from_date_raw, to_date_raw)
    start_rfc3339, end_rfc3339 = date_bounds_to_rfc3339(start_date, end_date)

    fetched = fetch_paginated_results(
        config,
        "v1/invoices",
        {
            "page": "1",
            "page_size": str(page_size),
            "date_start": start_rfc3339,
            "date_end": end_rfc3339,
        },
        timeout=timeout,
        max_pages=max_pages,
    )

    raw_invoices = fetched["results"]
    if not all(isinstance(invoice, dict) for invoice in raw_invoices):
        raise SiigoApiError("La respuesta de facturas contiene elementos con formato inesperado.")

    customer_hints: dict[str, dict[str, str]] = {}
    for invoice in raw_invoices:
        customer = invoice.get("customer")
        if not isinstance(customer, dict):
            continue
        customer_id = safe_text(customer.get("id"))
        if not customer_id:
            continue
        customer_hints[customer_id] = {
            "identification": safe_text(customer.get("identification")),
            "branch_office": safe_text(customer.get("branch_office")),
        }

    customers_map, lookup_failures = fetch_customer_details(config, customer_hints, timeout=timeout)

    invoice_rows: list[dict[str, Any]] = []
    customer_agg: dict[str, dict[str, Any]] = {}
    customer_month_agg: dict[str, dict[str, Any]] = {}
    total_sales = Decimal("0")
    total_balance = Decimal("0")
    invoice_count = 0
    annulled_count = 0

    for invoice in raw_invoices:
        if invoice.get("annulled") is True:
            annulled_count += 1
            if not include_annulled:
                continue

        invoice_date = safe_text(invoice.get("date"))
        if not invoice_date:
            continue
        customer = invoice.get("customer") if isinstance(invoice.get("customer"), dict) else {}
        customer_id = safe_text(customer.get("id")) or "unknown"
        customer_identification = safe_text(customer.get("identification"))
        customer_branch_office = safe_text(customer.get("branch_office"))
        customer_body = customers_map.get(customer_id)
        fallback_name = customer_identification or customer_id
        customer_name = customer_display_name(customer_body, fallback_name)

        fx = invoice_exchange_rate(invoice)
        currency_obj = invoice.get("currency") if isinstance(invoice.get("currency"), dict) else {}
        currency_code = safe_text(currency_obj.get("code")) or "COP"
        total_value = to_decimal(invoice.get("total")) * fx
        balance_value = to_decimal(invoice.get("balance")) * fx
        paid_value = total_value - balance_value
        month_key = invoice_date[:7]

        row = {
            "invoice_id": safe_text(invoice.get("id")),
            "invoice_name": safe_text(invoice.get("name")),
            "prefix": safe_text(invoice.get("prefix")),
            "number": safe_text(invoice.get("number")),
            "date": invoice_date,
            "month": month_key,
            "customer_id": customer_id,
            "customer_identification": customer_identification,
            "customer_branch_office": customer_branch_office,
            "customer_name": customer_name,
            "document_id": safe_text(invoice.get("document", {}).get("id"))
            if isinstance(invoice.get("document"), dict)
            else "",
            "total": money_to_float(total_value),
            "balance": money_to_float(balance_value),
            "paid": money_to_float(paid_value),
            "annulled": bool(invoice.get("annulled")),
            "seller": safe_text(invoice.get("seller")),
            "currency_code": currency_code,
        }
        invoice_rows.append(row)

        customer_key = f"{customer_id}|{customer_branch_office}"
        customer_entry = customer_agg.setdefault(
            customer_key,
            {
                "customer_id": customer_id,
                "customer_identification": customer_identification,
                "customer_branch_office": customer_branch_office,
                "customer_name": customer_name,
                "invoice_count": 0,
                "total_sales": Decimal("0"),
                "outstanding_balance": Decimal("0"),
                "paid_amount": Decimal("0"),
            },
        )
        customer_entry["invoice_count"] += 1
        customer_entry["total_sales"] += total_value
        customer_entry["outstanding_balance"] += balance_value
        customer_entry["paid_amount"] += paid_value

        month_customer_key = f"{month_key}|{customer_key}"
        month_entry = customer_month_agg.setdefault(
            month_customer_key,
            {
                "month": month_key,
                "customer_id": customer_id,
                "customer_identification": customer_identification,
                "customer_branch_office": customer_branch_office,
                "customer_name": customer_name,
                "invoice_count": 0,
                "total_sales": Decimal("0"),
                "outstanding_balance": Decimal("0"),
                "paid_amount": Decimal("0"),
            },
        )
        month_entry["invoice_count"] += 1
        month_entry["total_sales"] += total_value
        month_entry["outstanding_balance"] += balance_value
        month_entry["paid_amount"] += paid_value

        total_sales += total_value
        total_balance += balance_value
        invoice_count += 1

    # ── Notas crédito ─────────────────────────────────────────────────────────
    inv_fx_map: dict[str, Decimal] = {
        safe_text(inv.get("name")): invoice_exchange_rate(inv)
        for inv in raw_invoices
        if isinstance(inv, dict) and safe_text(inv.get("name"))
    }
    print("  Descargando notas crédito…")
    nc_list = _fetch_credit_notes_safe(config, start_date, end_date, timeout=timeout)
    print(f"  {len(nc_list)} notas crédito encontradas.")
    nc_customer_map, nc_customer_month_map = _build_nc_maps_by_customer(nc_list, inv_fx_map)
    total_credit_notes = sum(nc_customer_map.values())
    nc_rows = _build_nc_detail_rows(nc_list, inv_fx_map)

    # Mapa customer_key → info básica del cliente, extraída de las NCs crudas
    nc_customer_info: dict[str, dict[str, str]] = {}
    for _nc in nc_list:
        if _nc.get("annulled"):
            continue
        _c = _nc.get("customer") if isinstance(_nc.get("customer"), dict) else {}
        _cid    = safe_text(_c.get("id")) or "unknown"
        _branch = safe_text(_c.get("branch_office"))
        _key    = f"{_cid}|{_branch}"
        if _key not in nc_customer_info:
            _body = customers_map.get(_cid)
            _ident = safe_text(_c.get("identification"))
            nc_customer_info[_key] = {
                "customer_id": _cid,
                "customer_identification": _ident,
                "customer_branch_office": _branch,
                "customer_name": customer_display_name(_body, _ident or _cid),
            }

    customer_rows = [
        {
            **entry,
            "total_sales": money_to_float(entry["total_sales"]),
            "outstanding_balance": money_to_float(entry["outstanding_balance"]),
            "paid_amount": money_to_float(entry["paid_amount"]),
            "credit_notes_total": money_to_float(
                nc_customer_map.get(f"{entry['customer_id']}|{entry['customer_branch_office']}", Decimal("0"))
            ),
            "net_sales": money_to_float(
                entry["total_sales"]
                - nc_customer_map.get(f"{entry['customer_id']}|{entry['customer_branch_office']}", Decimal("0"))
            ),
        }
        for entry in customer_agg.values()
    ]
    # Clientes que solo tienen NCs (sin facturas en el período)
    for _key, _nc_total in nc_customer_map.items():
        if _key not in customer_agg and _key in nc_customer_info:
            _info = nc_customer_info[_key]
            customer_rows.append({
                **_info,
                "invoice_count": 0,
                "total_sales": 0.0,
                "outstanding_balance": 0.0,
                "paid_amount": 0.0,
                "credit_notes_total": money_to_float(_nc_total),
                "net_sales": money_to_float(-_nc_total),
            })
    customer_rows.sort(key=lambda item: (-item["total_sales"], item["customer_name"], item["customer_identification"]))

    customer_month_rows = [
        {
            **entry,
            "total_sales": money_to_float(entry["total_sales"]),
            "outstanding_balance": money_to_float(entry["outstanding_balance"]),
            "paid_amount": money_to_float(entry["paid_amount"]),
            "credit_notes_total": money_to_float(
                nc_customer_month_map.get(
                    f"{entry['month']}|{entry['customer_id']}|{entry['customer_branch_office']}", Decimal("0")
                )
            ),
            "net_sales": money_to_float(
                entry["total_sales"]
                - nc_customer_month_map.get(
                    f"{entry['month']}|{entry['customer_id']}|{entry['customer_branch_office']}", Decimal("0")
                )
            ),
        }
        for entry in customer_month_agg.values()
    ]
    customer_month_rows.sort(
        key=lambda item: (item["month"], -item["total_sales"], item["customer_name"], item["customer_identification"])
    )

    invoice_rows.sort(key=lambda item: (item["date"], item["invoice_name"], item["customer_name"]))

    return {
        "summary": {
            "report_type": "sales_by_customer",
            "from_date": start_date.isoformat(),
            "to_date": end_date.isoformat(),
            "date_start_rfc3339": start_rfc3339,
            "date_end_rfc3339": end_rfc3339,
            "months_requested": months,
            "include_annulled": include_annulled,
            "pages_fetched": fetched["pages_fetched"],
            "raw_invoices": len(raw_invoices),
            "report_invoices": invoice_count,
            "annulled_invoices_in_range": annulled_count,
            "customers": len(customer_rows),
            "total_sales": money_to_float(total_sales),
            "credit_notes_total": money_to_float(total_credit_notes),
            "net_sales": money_to_float(total_sales - total_credit_notes),
            "outstanding_balance": money_to_float(total_balance),
            "reported_total_results": fetched["reported_total_results"],
            "customer_lookup_failures": len(lookup_failures),
            "credit_notes_count": len(nc_list),
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        },
        "customer_lookup_failures": lookup_failures,
        "by_customer": customer_rows,
        "by_customer_month": customer_month_rows,
        "invoices": invoice_rows,
        "credit_notes": nc_rows,
    }


def build_billing_report(
    config: Config,
    months: int,
    from_date_raw: str | None,
    to_date_raw: str | None,
    include_annulled: bool,
    page_size: int,
    timeout: int = 30,
    max_pages: int | None = None,
) -> dict[str, Any]:
    start_date, end_date = resolve_report_range(months, from_date_raw, to_date_raw)
    start_rfc3339, end_rfc3339 = date_bounds_to_rfc3339(start_date, end_date)

    fetched = fetch_paginated_results(
        config,
        "v1/invoices",
        {
            "page": "1",
            "page_size": str(page_size),
            "date_start": start_rfc3339,
            "date_end": end_rfc3339,
        },
        timeout=timeout,
        max_pages=max_pages,
    )

    raw_invoices = fetched["results"]
    if not all(isinstance(invoice, dict) for invoice in raw_invoices):
        raise SiigoApiError("La respuesta de facturas contiene elementos con formato inesperado.")

    customer_hints: dict[str, dict[str, str]] = {}
    for invoice in raw_invoices:
        customer = invoice.get("customer")
        if not isinstance(customer, dict):
            continue
        customer_id = safe_text(customer.get("id"))
        if customer_id:
            customer_hints[customer_id] = {
                "identification": safe_text(customer.get("identification")),
                "branch_office": safe_text(customer.get("branch_office")),
            }

    customers_map, customer_lookup_failures = fetch_customer_details(
        config,
        customer_hints,
        timeout=timeout,
    )
    users_map, users_stats = fetch_users_map(config, timeout=timeout)

    invoice_rows: list[dict[str, Any]] = []
    customer_agg: dict[str, dict[str, Any]] = {}
    seller_agg: dict[str, dict[str, Any]] = {}
    customer_seller_agg: dict[str, dict[str, Any]] = {}
    day_agg: dict[str, dict[str, Any]] = {}

    total_sales = Decimal("0")
    total_balance = Decimal("0")
    invoice_count = 0
    annulled_count = 0
    unresolved_sellers: set[str] = set()

    for invoice in raw_invoices:
        if invoice.get("annulled") is True:
            annulled_count += 1
            if not include_annulled:
                continue

        invoice_date = safe_text(invoice.get("date"))
        if not invoice_date:
            continue
        month_key = invoice_date[:7]

        customer = invoice.get("customer") if isinstance(invoice.get("customer"), dict) else {}
        customer_id = safe_text(customer.get("id")) or "unknown"
        customer_identification = safe_text(customer.get("identification"))
        customer_branch_office = safe_text(customer.get("branch_office"))
        customer_body = customers_map.get(customer_id)
        customer_fallback_name = customer_identification or customer_id or "Sin cliente"
        customer_name = customer_display_name(customer_body, customer_fallback_name)

        seller_id = safe_text(invoice.get("seller"))
        seller_body = users_map.get(seller_id)
        if seller_id and seller_body is None:
            unresolved_sellers.add(seller_id)
        seller_name = user_display_name(
            seller_body,
            seller_id or "Sin vendedor",
        )
        seller_email = safe_text(seller_body.get("email")) if isinstance(seller_body, dict) else ""
        seller_identification = (
            safe_text(seller_body.get("identification")) if isinstance(seller_body, dict) else ""
        )

        fx = invoice_exchange_rate(invoice)
        currency_obj = invoice.get("currency") if isinstance(invoice.get("currency"), dict) else {}
        currency_code = safe_text(currency_obj.get("code")) or "COP"
        total_value = to_decimal(invoice.get("total")) * fx
        balance_value = to_decimal(invoice.get("balance")) * fx
        paid_value = total_value - balance_value

        row = {
            "date": invoice_date,
            "month": month_key,
            "invoice_name": safe_text(invoice.get("name")),
            "prefix": safe_text(invoice.get("prefix")),
            "number": safe_text(invoice.get("number")),
            "invoice_id": safe_text(invoice.get("id")),
            "document_id": safe_text(invoice.get("document", {}).get("id"))
            if isinstance(invoice.get("document"), dict)
            else "",
            "customer_name": customer_name,
            "customer_identification": customer_identification,
            "customer_branch_office": customer_branch_office,
            "customer_id": customer_id,
            "seller_id": seller_id,
            "seller_name": seller_name,
            "seller_identification": seller_identification,
            "seller_email": seller_email,
            "total": money_to_float(total_value),
            "paid": money_to_float(paid_value),
            "balance": money_to_float(balance_value),
            "annulled": bool(invoice.get("annulled")),
            "currency_code": currency_code,
        }
        invoice_rows.append(row)

        customer_key = f"{customer_id}|{customer_branch_office}"
        customer_entry = customer_agg.setdefault(
            customer_key,
            {
                "customer_id": customer_id,
                "customer_identification": customer_identification,
                "customer_branch_office": customer_branch_office,
                "customer_name": customer_name,
                "invoice_count": 0,
                "total_sales": Decimal("0"),
                "paid_amount": Decimal("0"),
                "outstanding_balance": Decimal("0"),
                "_sellers": set(),
            },
        )
        customer_entry["invoice_count"] += 1
        customer_entry["total_sales"] += total_value
        customer_entry["paid_amount"] += paid_value
        customer_entry["outstanding_balance"] += balance_value
        customer_entry["_sellers"].add(seller_id or "Sin vendedor")

        seller_key = seller_id or "Sin vendedor"
        seller_entry = seller_agg.setdefault(
            seller_key,
            {
                "seller_id": seller_id,
                "seller_name": seller_name,
                "seller_identification": seller_identification,
                "seller_email": seller_email,
                "invoice_count": 0,
                "total_sales": Decimal("0"),
                "paid_amount": Decimal("0"),
                "outstanding_balance": Decimal("0"),
                "_customers": set(),
            },
        )
        seller_entry["invoice_count"] += 1
        seller_entry["total_sales"] += total_value
        seller_entry["paid_amount"] += paid_value
        seller_entry["outstanding_balance"] += balance_value
        seller_entry["_customers"].add(customer_key)

        customer_seller_key = f"{customer_key}|{seller_key}"
        customer_seller_entry = customer_seller_agg.setdefault(
            customer_seller_key,
            {
                "customer_id": customer_id,
                "customer_identification": customer_identification,
                "customer_branch_office": customer_branch_office,
                "customer_name": customer_name,
                "seller_id": seller_id,
                "seller_name": seller_name,
                "seller_identification": seller_identification,
                "seller_email": seller_email,
                "invoice_count": 0,
                "total_sales": Decimal("0"),
                "paid_amount": Decimal("0"),
                "outstanding_balance": Decimal("0"),
            },
        )
        customer_seller_entry["invoice_count"] += 1
        customer_seller_entry["total_sales"] += total_value
        customer_seller_entry["paid_amount"] += paid_value
        customer_seller_entry["outstanding_balance"] += balance_value

        day_entry = day_agg.setdefault(
            invoice_date,
            {
                "date": invoice_date,
                "month": month_key,
                "invoice_count": 0,
                "total_sales": Decimal("0"),
                "paid_amount": Decimal("0"),
                "outstanding_balance": Decimal("0"),
                "_customers": set(),
                "_sellers": set(),
            },
        )
        day_entry["invoice_count"] += 1
        day_entry["total_sales"] += total_value
        day_entry["paid_amount"] += paid_value
        day_entry["outstanding_balance"] += balance_value
        day_entry["_customers"].add(customer_key)
        day_entry["_sellers"].add(seller_key)

        total_sales += total_value
        total_balance += balance_value
        invoice_count += 1

    # ── Notas crédito ─────────────────────────────────────────────────────────
    inv_fx_map: dict[str, Decimal] = {
        safe_text(inv.get("name")): invoice_exchange_rate(inv)
        for inv in raw_invoices
        if isinstance(inv, dict) and safe_text(inv.get("name"))
    }
    print("  Descargando notas crédito…")
    nc_list = _fetch_credit_notes_safe(config, start_date, end_date, timeout=timeout)
    print(f"  {len(nc_list)} notas crédito encontradas.")
    nc_customer_map, _nc_cm_map = _build_nc_maps_by_customer(nc_list, inv_fx_map)
    nc_day_map: dict[str, Decimal] = {}
    for _nc in nc_list:
        if _nc.get("annulled"):
            continue
        _nd = safe_text(_nc.get("date"))
        if _nd:
            _nc_fx_rate = _nc_fx(_nc, inv_fx_map)
            nc_day_map[_nd] = nc_day_map.get(_nd, Decimal("0")) + to_decimal(_nc.get("total")) * _nc_fx_rate
    total_credit_notes = sum(nc_customer_map.values())
    nc_rows = _build_nc_detail_rows(nc_list, inv_fx_map)

    customer_rows = [
        {
            "customer_id": entry["customer_id"],
            "customer_identification": entry["customer_identification"],
            "customer_branch_office": entry["customer_branch_office"],
            "customer_name": entry["customer_name"],
            "invoice_count": entry["invoice_count"],
            "seller_count": len(entry["_sellers"]),
            "total_sales": money_to_float(entry["total_sales"]),
            "paid_amount": money_to_float(entry["paid_amount"]),
            "outstanding_balance": money_to_float(entry["outstanding_balance"]),
            "avg_invoice_total": average_amount(entry["total_sales"], entry["invoice_count"]),
            "credit_notes_total": money_to_float(
                nc_customer_map.get(f"{entry['customer_id']}|{entry['customer_branch_office']}", Decimal("0"))
            ),
            "net_sales": money_to_float(
                entry["total_sales"]
                - nc_customer_map.get(f"{entry['customer_id']}|{entry['customer_branch_office']}", Decimal("0"))
            ),
        }
        for entry in customer_agg.values()
    ]
    customer_rows.sort(key=lambda item: (-item["total_sales"], item["customer_name"], item["customer_identification"]))

    seller_rows = [
        {
            "seller_id": entry["seller_id"],
            "seller_name": entry["seller_name"],
            "seller_identification": entry["seller_identification"],
            "seller_email": entry["seller_email"],
            "invoice_count": entry["invoice_count"],
            "customer_count": len(entry["_customers"]),
            "total_sales": money_to_float(entry["total_sales"]),
            "paid_amount": money_to_float(entry["paid_amount"]),
            "outstanding_balance": money_to_float(entry["outstanding_balance"]),
            "avg_invoice_total": average_amount(entry["total_sales"], entry["invoice_count"]),
        }
        for entry in seller_agg.values()
    ]
    seller_rows.sort(key=lambda item: (-item["total_sales"], item["seller_name"], safe_text(item["seller_id"])))

    customer_seller_rows = [
        {
            "customer_id": entry["customer_id"],
            "customer_identification": entry["customer_identification"],
            "customer_branch_office": entry["customer_branch_office"],
            "customer_name": entry["customer_name"],
            "seller_id": entry["seller_id"],
            "seller_name": entry["seller_name"],
            "seller_identification": entry["seller_identification"],
            "seller_email": entry["seller_email"],
            "invoice_count": entry["invoice_count"],
            "total_sales": money_to_float(entry["total_sales"]),
            "paid_amount": money_to_float(entry["paid_amount"]),
            "outstanding_balance": money_to_float(entry["outstanding_balance"]),
            "avg_invoice_total": average_amount(entry["total_sales"], entry["invoice_count"]),
        }
        for entry in customer_seller_agg.values()
    ]
    customer_seller_rows.sort(
        key=lambda item: (-item["total_sales"], item["customer_name"], item["seller_name"])
    )

    day_rows = [
        {
            "date": entry["date"],
            "month": entry["month"],
            "invoice_count": entry["invoice_count"],
            "customer_count": len(entry["_customers"]),
            "seller_count": len(entry["_sellers"]),
            "total_sales": money_to_float(entry["total_sales"]),
            "paid_amount": money_to_float(entry["paid_amount"]),
            "outstanding_balance": money_to_float(entry["outstanding_balance"]),
            "avg_invoice_total": average_amount(entry["total_sales"], entry["invoice_count"]),
            "credit_notes_total": money_to_float(nc_day_map.get(entry["date"], Decimal("0"))),
            "net_sales": money_to_float(
                entry["total_sales"] - nc_day_map.get(entry["date"], Decimal("0"))
            ),
        }
        for entry in day_agg.values()
    ]
    day_rows.sort(key=lambda item: item["date"])

    invoice_rows.sort(
        key=lambda item: (
            item["date"],
            item["seller_name"],
            item["customer_name"],
            item["invoice_name"],
        )
    )

    return {
        "summary": {
            "report_type": "billing_detail",
            "from_date": start_date.isoformat(),
            "to_date": end_date.isoformat(),
            "date_start_rfc3339": start_rfc3339,
            "date_end_rfc3339": end_rfc3339,
            "months_requested": months,
            "include_annulled": include_annulled,
            "pages_fetched": fetched["pages_fetched"],
            "raw_invoices": len(raw_invoices),
            "report_invoices": invoice_count,
            "annulled_invoices_in_range": annulled_count,
            "customers": len(customer_rows),
            "sellers": len(seller_rows),
            "customer_seller_pairs": len(customer_seller_rows),
            "total_sales": money_to_float(total_sales),
            "credit_notes_total": money_to_float(total_credit_notes),
            "net_sales": money_to_float(total_sales - total_credit_notes),
            "outstanding_balance": money_to_float(total_balance),
            "reported_total_results": fetched["reported_total_results"],
            "customer_lookup_failures": len(customer_lookup_failures),
            "users_pages_fetched": users_stats["pages_fetched"],
            "users_reported_total": users_stats["reported_total_results"],
            "unresolved_sellers": len(unresolved_sellers),
            "credit_notes_count": len(nc_list),
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        },
        "customer_lookup_failures": customer_lookup_failures,
        "unresolved_seller_ids": sorted(unresolved_sellers),
        "by_customer": customer_rows,
        "by_seller": seller_rows,
        "by_customer_seller": customer_seller_rows,
        "by_day": day_rows,
        "invoices": invoice_rows,
        "credit_notes": nc_rows,
    }


def write_sales_workbook(output_path: Path, report: dict[str, Any]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    append_dict_sheet(
        workbook,
        "Resumen",
        build_summary_rows(report["summary"]),
        table_name="ResumenVentas",
    )
    append_dict_sheet(
        workbook,
        "Facturas",
        report["invoices"],
        table_name="FacturasDetalle",
        numeric_columns={"total", "balance", "paid"},
        date_columns={"date"},
    )
    append_dict_sheet(
        workbook,
        "PorCliente",
        report["by_customer"],
        table_name="VentasPorCliente",
        numeric_columns={
            "invoice_count", "total_sales", "outstanding_balance",
            "paid_amount", "credit_notes_total", "net_sales",
        },
    )
    append_dict_sheet(
        workbook,
        "PorClienteMes",
        report["by_customer_month"],
        table_name="VentasPorClienteMes",
        numeric_columns={
            "invoice_count", "total_sales", "outstanding_balance",
            "paid_amount", "credit_notes_total", "net_sales",
        },
    )
    if report.get("credit_notes"):
        append_dict_sheet(
            workbook,
            "NotasCredito",
            report["credit_notes"],
            table_name="NotasCreditoVentas",
            numeric_columns={"total"},
            date_columns={"date"},
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def write_cartera_workbook(output_path: Path, report: dict[str, Any]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    append_dict_sheet(
        workbook,
        "Resumen",
        build_summary_rows(report["summary"]),
        table_name="ResumenCartera",
    )
    append_dict_sheet(
        workbook,
        "PorAntiguedad",
        report["by_bucket"],
        table_name="CarteraPorAntiguedad",
        numeric_columns={"total_balance"},
    )
    append_dict_sheet(
        workbook,
        "PorCliente",
        report["by_customer"],
        table_name="CarteraPorCliente",
        numeric_columns={
            "invoice_count",
            "total_balance",
            "current_30",
            "overdue_31_60",
            "overdue_61_90",
            "overdue_91_plus",
        },
    )
    append_dict_sheet(
        workbook,
        "FacturasPendientes",
        report["invoices"],
        table_name="FacturasPendientes",
        numeric_columns={"total", "balance", "overdue_days"},
        date_columns={"date", "due_date"},
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def write_billing_workbook(output_path: Path, report: dict[str, Any]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    append_dict_sheet(
        workbook,
        "Resumen",
        build_summary_rows(report["summary"]),
        table_name="ResumenFacturacion",
    )
    append_dict_sheet(
        workbook,
        "Facturas",
        report["invoices"],
        table_name="FacturasDetalle",
        numeric_columns={"total", "paid", "balance"},
        date_columns={"date"},
    )
    append_dict_sheet(
        workbook,
        "PorCliente",
        report["by_customer"],
        table_name="FacturacionPorCliente",
        numeric_columns={
            "invoice_count",
            "seller_count",
            "total_sales",
            "avg_invoice_total",
            "paid_amount",
            "outstanding_balance",
            "credit_notes_total",
            "net_sales",
        },
    )
    append_dict_sheet(
        workbook,
        "PorVendedor",
        report["by_seller"],
        table_name="FacturacionPorVendedor",
        numeric_columns={
            "invoice_count",
            "customer_count",
            "total_sales",
            "avg_invoice_total",
            "paid_amount",
            "outstanding_balance",
        },
    )
    append_dict_sheet(
        workbook,
        "ClienteVendedor",
        report["by_customer_seller"],
        table_name="FacturacionClienteVendedor",
        numeric_columns={
            "invoice_count",
            "total_sales",
            "avg_invoice_total",
            "paid_amount",
            "outstanding_balance",
        },
    )
    append_dict_sheet(
        workbook,
        "PorDia",
        report["by_day"],
        table_name="FacturacionPorDia",
        numeric_columns={
            "invoice_count",
            "customer_count",
            "seller_count",
            "total_sales",
            "avg_invoice_total",
            "paid_amount",
            "outstanding_balance",
            "credit_notes_total",
            "net_sales",
        },
        date_columns={"date"},
    )
    if report.get("credit_notes"):
        append_dict_sheet(
            workbook,
            "NotasCredito",
            report["credit_notes"],
            table_name="NotasCredito",
            numeric_columns={"total"},
            date_columns={"date"},
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def write_billing_report_files(output_path: Path, report: dict[str, Any]) -> dict[str, Path]:
    write_output(output_path, report)

    base_name = output_path.stem
    output_dir = output_path.parent
    informes_dir = output_dir / "informes"
    informes_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = informes_dir / f"{base_name}.xlsx"
    customer_csv = output_dir / f"{base_name}_by_customer.csv"
    seller_csv = output_dir / f"{base_name}_by_seller.csv"
    customer_seller_csv = output_dir / f"{base_name}_by_customer_seller.csv"
    day_csv = output_dir / f"{base_name}_by_day.csv"
    invoice_csv = output_dir / f"{base_name}_invoices.csv"

    write_billing_workbook(workbook_path, report)

    write_csv(
        customer_csv,
        [
            "customer_name",
            "customer_identification",
            "customer_branch_office",
            "customer_id",
            "invoice_count",
            "seller_count",
            "total_sales",
            "credit_notes_total",
            "net_sales",
            "avg_invoice_total",
            "paid_amount",
            "outstanding_balance",
        ],
        report["by_customer"],
    )
    write_csv(
        seller_csv,
        [
            "seller_name",
            "seller_id",
            "seller_identification",
            "seller_email",
            "invoice_count",
            "customer_count",
            "total_sales",
            "avg_invoice_total",
            "paid_amount",
            "outstanding_balance",
        ],
        report["by_seller"],
    )
    write_csv(
        customer_seller_csv,
        [
            "customer_name",
            "customer_identification",
            "customer_branch_office",
            "customer_id",
            "seller_name",
            "seller_id",
            "seller_identification",
            "seller_email",
            "invoice_count",
            "total_sales",
            "avg_invoice_total",
            "paid_amount",
            "outstanding_balance",
        ],
        report["by_customer_seller"],
    )
    write_csv(
        day_csv,
        [
            "date",
            "month",
            "invoice_count",
            "customer_count",
            "seller_count",
            "total_sales",
            "credit_notes_total",
            "net_sales",
            "avg_invoice_total",
            "paid_amount",
            "outstanding_balance",
        ],
        report["by_day"],
    )
    write_csv(
        invoice_csv,
        [
            "date",
            "month",
            "invoice_name",
            "prefix",
            "number",
            "invoice_id",
            "document_id",
            "customer_name",
            "customer_identification",
            "customer_branch_office",
            "customer_id",
            "seller_name",
            "seller_id",
            "seller_identification",
            "seller_email",
            "total",
            "paid",
            "balance",
            "annulled",
            "currency_code",
        ],
        report["invoices"],
    )

    nc_csv = output_dir / f"{base_name}_credit_notes.csv"
    if report.get("credit_notes"):
        write_csv(
            nc_csv,
            ["date", "month", "nc_name", "nc_id", "customer_id",
             "customer_identification", "customer_branch_office", "linked_invoice", "total"],
            report["credit_notes"],
        )

    return {
        "json": output_path,
        "xlsx": workbook_path,
        "by_customer_csv": customer_csv,
        "by_seller_csv": seller_csv,
        "by_customer_seller_csv": customer_seller_csv,
        "by_day_csv": day_csv,
        "invoices_csv": invoice_csv,
        "credit_notes_csv": nc_csv,
    }


def write_sales_report_files(output_path: Path, report: dict[str, Any]) -> dict[str, Path]:
    write_output(output_path, report)

    base_name = output_path.stem
    output_dir = output_path.parent
    informes_dir = output_dir / "informes"
    informes_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = informes_dir / f"{base_name}.xlsx"
    customer_csv = output_dir / f"{base_name}_by_customer.csv"
    customer_month_csv = output_dir / f"{base_name}_by_customer_month.csv"
    invoice_csv = output_dir / f"{base_name}_invoices.csv"

    write_sales_workbook(workbook_path, report)

    write_csv(
        customer_csv,
        [
            "customer_name",
            "customer_identification",
            "customer_branch_office",
            "customer_id",
            "invoice_count",
            "total_sales",
            "credit_notes_total",
            "net_sales",
            "paid_amount",
            "outstanding_balance",
        ],
        report["by_customer"],
    )
    write_csv(
        customer_month_csv,
        [
            "month",
            "customer_name",
            "customer_identification",
            "customer_branch_office",
            "customer_id",
            "invoice_count",
            "total_sales",
            "credit_notes_total",
            "net_sales",
            "paid_amount",
            "outstanding_balance",
        ],
        report["by_customer_month"],
    )
    write_csv(
        invoice_csv,
        [
            "date",
            "month",
            "invoice_name",
            "prefix",
            "number",
            "invoice_id",
            "document_id",
            "customer_name",
            "customer_identification",
            "customer_branch_office",
            "customer_id",
            "total",
            "paid",
            "balance",
            "annulled",
            "seller",
            "currency_code",
        ],
        report["invoices"],
    )

    nc_csv = output_dir / f"{base_name}_credit_notes.csv"
    if report.get("credit_notes"):
        write_csv(
            nc_csv,
            ["date", "month", "nc_name", "nc_id", "customer_id",
             "customer_identification", "customer_branch_office", "linked_invoice", "total"],
            report["credit_notes"],
        )

    return {
        "json": output_path,
        "xlsx": workbook_path,
        "by_customer_csv": customer_csv,
        "by_customer_month_csv": customer_month_csv,
        "invoices_csv": invoice_csv,
        "credit_notes_csv": nc_csv,
    }


def days_overdue(invoice_date_str: str) -> int:
    if not invoice_date_str:
        return 0
    try:
        invoice_dt = date.fromisoformat(invoice_date_str)
        return (date.today() - invoice_dt).days
    except ValueError:
        return 0


def aging_bucket(days: int) -> str:
    if days <= 30:
        return "Corriente"
    elif days <= 60:
        return "31-60 días"
    elif days <= 90:
        return "61-90 días"
    else:
        return "Más de 90 días"


def build_cartera_report(
    config: Config,
    months: int,
    from_date_raw: str | None,
    to_date_raw: str | None,
    page_size: int,
    timeout: int = 30,
    max_pages: int | None = None,
) -> dict[str, Any]:
    start_date, end_date = resolve_report_range(months, from_date_raw, to_date_raw)
    start_rfc3339, end_rfc3339 = date_bounds_to_rfc3339(start_date, end_date)

    fetched = fetch_paginated_results(
        config,
        "v1/invoices",
        {
            "page": "1",
            "page_size": str(page_size),
            "date_start": start_rfc3339,
            "date_end": end_rfc3339,
        },
        timeout=timeout,
        max_pages=max_pages,
    )

    raw_invoices = fetched["results"]
    if not all(isinstance(invoice, dict) for invoice in raw_invoices):
        raise SiigoApiError("La respuesta de facturas contiene elementos con formato inesperado.")

    customer_hints: dict[str, dict[str, str]] = {}
    for invoice in raw_invoices:
        customer = invoice.get("customer")
        if not isinstance(customer, dict):
            continue
        customer_id = safe_text(customer.get("id"))
        if customer_id:
            customer_hints[customer_id] = {
                "identification": safe_text(customer.get("identification")),
                "branch_office": safe_text(customer.get("branch_office")),
            }

    customers_map, customer_lookup_failures = fetch_customer_details(
        config,
        customer_hints,
        timeout=timeout,
    )

    today = date.today()
    invoice_rows: list[dict[str, Any]] = []
    customer_agg: dict[str, dict[str, Any]] = {}
    bucket_agg: dict[str, Decimal] = {
        "Corriente": Decimal("0"),
        "31-60 días": Decimal("0"),
        "61-90 días": Decimal("0"),
        "Más de 90 días": Decimal("0"),
    }

    total_balance = Decimal("0")
    invoice_count = 0
    annulled_count = 0

    for invoice in raw_invoices:
        if invoice.get("annulled") is True:
            annulled_count += 1
            continue

        invoice_date_str = safe_text(invoice.get("date"))
        if not invoice_date_str:
            continue

        fx = invoice_exchange_rate(invoice)
        balance_value = to_decimal(invoice.get("balance")) * fx
        if balance_value <= 0:
            continue

        customer = invoice.get("customer") if isinstance(invoice.get("customer"), dict) else {}
        customer_id = safe_text(customer.get("id")) or "unknown"
        customer_identification = safe_text(customer.get("identification"))
        customer_branch_office = safe_text(customer.get("branch_office"))
        customer_body = customers_map.get(customer_id)
        fallback_name = customer_identification or customer_id
        customer_name = customer_display_name(customer_body, fallback_name)

        overdue_days = days_overdue(invoice_date_str)
        bucket = aging_bucket(overdue_days)
        total_value = to_decimal(invoice.get("total")) * fx

        invoice_rows.append({
            "invoice_id": safe_text(invoice.get("id")),
            "invoice_name": safe_text(invoice.get("name")),
            "prefix": safe_text(invoice.get("prefix")),
            "number": safe_text(invoice.get("number")),
            "date": invoice_date_str,
            "due_date": safe_text(invoice.get("due_date")),
            "overdue_days": overdue_days,
            "aging_bucket": bucket,
            "customer_id": customer_id,
            "customer_identification": customer_identification,
            "customer_branch_office": customer_branch_office,
            "customer_name": customer_name,
            "total": money_to_float(total_value),
            "balance": money_to_float(balance_value),
            "currency_code": safe_text(invoice.get("currency", {}).get("code"))
            if isinstance(invoice.get("currency"), dict)
            else "",
        })

        bucket_agg[bucket] += balance_value

        customer_key = f"{customer_id}|{customer_branch_office}"
        customer_entry = customer_agg.setdefault(
            customer_key,
            {
                "customer_id": customer_id,
                "customer_identification": customer_identification,
                "customer_branch_office": customer_branch_office,
                "customer_name": customer_name,
                "invoice_count": 0,
                "total_balance": Decimal("0"),
                "Corriente": Decimal("0"),
                "31-60 días": Decimal("0"),
                "61-90 días": Decimal("0"),
                "Más de 90 días": Decimal("0"),
            },
        )
        customer_entry["invoice_count"] += 1
        customer_entry["total_balance"] += balance_value
        customer_entry[bucket] += balance_value

        total_balance += balance_value
        invoice_count += 1

    customer_rows = [
        {
            "customer_id": entry["customer_id"],
            "customer_identification": entry["customer_identification"],
            "customer_branch_office": entry["customer_branch_office"],
            "customer_name": entry["customer_name"],
            "invoice_count": entry["invoice_count"],
            "total_balance": money_to_float(entry["total_balance"]),
            "current_30": money_to_float(entry["Corriente"]),
            "overdue_31_60": money_to_float(entry["31-60 días"]),
            "overdue_61_90": money_to_float(entry["61-90 días"]),
            "overdue_91_plus": money_to_float(entry["Más de 90 días"]),
        }
        for entry in customer_agg.values()
    ]
    customer_rows.sort(key=lambda item: (-item["total_balance"], item["customer_name"]))

    bucket_rows = [
        {"bucket": bucket_name, "total_balance": money_to_float(amount)}
        for bucket_name, amount in sorted(
            bucket_agg.items(),
            key=lambda item: {"Corriente": 0, "31-60 días": 1, "61-90 días": 2, "Más de 90 días": 3}.get(item[0], 99),
        )
    ]

    invoice_rows.sort(key=lambda item: (item["aging_bucket"], item["customer_name"], item["date"]))

    return {
        "summary": {
            "report_type": "cartera",
            "generated_at": today.isoformat(),
            "from_date": start_date.isoformat(),
            "to_date": end_date.isoformat(),
            "pages_fetched": fetched["pages_fetched"],
            "raw_invoices_in_range": len(raw_invoices),
            "annulled_excluded": annulled_count,
            "invoices_with_balance": invoice_count,
            "customers_with_debt": len(customer_rows),
            "total_balance": money_to_float(total_balance),
            "reported_total_results": fetched["reported_total_results"],
            "customer_lookup_failures": len(customer_lookup_failures),
        },
        "customer_lookup_failures": customer_lookup_failures,
        "by_bucket": bucket_rows,
        "by_customer": customer_rows,
        "invoices": invoice_rows,
    }


def write_cartera_report_files(output_path: Path, report: dict[str, Any]) -> dict[str, Path]:
    write_output(output_path, report)

    base_name = output_path.stem
    output_dir = output_path.parent
    informes_dir = output_dir / "informes"
    informes_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = informes_dir / f"{base_name}.xlsx"
    customer_csv = output_dir / f"{base_name}_by_customer.csv"
    invoice_csv = output_dir / f"{base_name}_invoices.csv"

    write_cartera_workbook(workbook_path, report)

    write_csv(
        customer_csv,
        [
            "customer_name",
            "customer_identification",
            "customer_branch_office",
            "customer_id",
            "invoice_count",
            "total_balance",
            "current_30",
            "overdue_31_60",
            "overdue_61_90",
            "overdue_91_plus",
        ],
        report["by_customer"],
    )
    write_csv(
        invoice_csv,
        [
            "invoice_id",
            "invoice_name",
            "prefix",
            "number",
            "date",
            "due_date",
            "overdue_days",
            "aging_bucket",
            "customer_name",
            "customer_identification",
            "customer_branch_office",
            "customer_id",
            "total",
            "balance",
            "currency_code",
        ],
        report["invoices"],
    )

    return {
        "json": output_path,
        "xlsx": workbook_path,
        "by_customer_csv": customer_csv,
        "invoices_csv": invoice_csv,
    }


def mask_token(token: str) -> str:
    if len(token) <= 8:
        return "*" * len(token)
    return f"{token[:4]}...{token[-4:]}"


def write_output(output_path: Path, body: Any) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(body, (dict, list)):
        content = json.dumps(body, indent=2, ensure_ascii=False)
    elif body is None:
        content = ""
    else:
        content = str(body)
    output_path.write_text(content, encoding="utf-8")


def print_body(body: Any) -> None:
    if isinstance(body, (dict, list)):
        print(json.dumps(body, indent=2, ensure_ascii=False))
        return
    if body is None:
        print("(sin contenido)")
        return
    print(body)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Cliente de pruebas de solo lectura para Siigo Nube."
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="Tiempo máximo por request en segundos. Default: 30",
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    subparsers.add_parser(
        "auth-test",
        help="Valida credenciales y obtiene un token sin mostrarlo completo.",
    )

    get_parser = subparsers.add_parser(
        "get",
        help="Ejecuta una consulta GET contra un endpoint v1/* de Siigo.",
    )
    get_parser.add_argument(
        "path",
        help="Path relativo del endpoint. Ejemplo: v1/customers",
    )
    get_parser.add_argument(
        "--param",
        action="append",
        default=[],
        help="Parámetro query en formato clave=valor. Se puede repetir.",
    )
    get_parser.add_argument(
        "--output",
        help="Ruta opcional para guardar la respuesta.",
    )

    export_parser = subparsers.add_parser(
        "export",
        help="Exporta todas las páginas de un endpoint tipo lista usando GET.",
    )
    export_parser.add_argument(
        "path",
        help="Path relativo del endpoint. Ejemplo: v1/products",
    )
    export_parser.add_argument(
        "--param",
        action="append",
        default=[],
        help="Parámetro query en formato clave=valor. Se puede repetir.",
    )
    export_parser.add_argument(
        "--output",
        required=True,
        help="Ruta donde se guardará el JSON exportado.",
    )
    export_parser.add_argument(
        "--max-pages",
        type=int,
        help="Límite opcional de páginas a descargar para pruebas.",
    )

    sales_report_parser = subparsers.add_parser(
        "sales-report",
        help="Construye un histórico de ventas por cliente a partir de facturas.",
    )
    sales_report_parser.add_argument(
        "--months",
        type=int,
        default=2,
        help="Cantidad de meses hacia atrás si no se define --from-date. Default: 2",
    )
    sales_report_parser.add_argument(
        "--from-date",
        help="Fecha inicial en formato YYYY-MM-DD.",
    )
    sales_report_parser.add_argument(
        "--to-date",
        help="Fecha final en formato YYYY-MM-DD. Por defecto usa la fecha de hoy.",
    )
    sales_report_parser.add_argument(
        "--page-size",
        type=int,
        default=25,
        help="Cantidad solicitada por página a Siigo. Default: 25",
    )
    sales_report_parser.add_argument(
        "--max-pages",
        type=int,
        help="Límite opcional de páginas a descargar para pruebas.",
    )
    sales_report_parser.add_argument(
        "--include-annulled",
        action="store_true",
        help="Incluye facturas anuladas en el reporte.",
    )
    sales_report_parser.add_argument(
        "--output",
        required=True,
        help="Ruta base del JSON de salida. Se generarán CSVs auxiliares al lado.",
    )

    billing_report_parser = subparsers.add_parser(
        "billing-report",
        help="Construye un reporte detallado de facturación por cliente y vendedor.",
    )
    billing_report_parser.add_argument(
        "--months",
        type=int,
        default=1,
        help="Cantidad de meses hacia atrás si no se define --from-date. Default: 1",
    )
    billing_report_parser.add_argument(
        "--from-date",
        help="Fecha inicial en formato YYYY-MM-DD.",
    )
    billing_report_parser.add_argument(
        "--to-date",
        help="Fecha final en formato YYYY-MM-DD. Por defecto usa la fecha de hoy.",
    )
    billing_report_parser.add_argument(
        "--page-size",
        type=int,
        default=25,
        help="Cantidad solicitada por página a Siigo. Default: 25",
    )
    billing_report_parser.add_argument(
        "--max-pages",
        type=int,
        help="Límite opcional de páginas a descargar para pruebas.",
    )
    billing_report_parser.add_argument(
        "--include-annulled",
        action="store_true",
        help="Incluye facturas anuladas en el reporte.",
    )
    billing_report_parser.add_argument(
        "--output",
        required=True,
        help="Ruta base del JSON de salida. Se generarán CSVs auxiliares al lado.",
    )

    cartera_parser = subparsers.add_parser(
        "cartera-report",
        help="Construye un reporte de cartera (cuentas por cobrar) por cliente y antigüedad.",
    )
    cartera_parser.add_argument(
        "--months",
        type=int,
        default=12,
        help="Cantidad de meses hacia atrás para buscar facturas. Default: 12",
    )
    cartera_parser.add_argument(
        "--from-date",
        help="Fecha inicial en formato YYYY-MM-DD.",
    )
    cartera_parser.add_argument(
        "--to-date",
        help="Fecha final en formato YYYY-MM-DD. Por defecto usa la fecha de hoy.",
    )
    cartera_parser.add_argument(
        "--page-size",
        type=int,
        default=100,
        help="Cantidad solicitada por página a Siigo. Default: 100",
    )
    cartera_parser.add_argument(
        "--max-pages",
        type=int,
        help="Límite opcional de páginas a descargar para pruebas.",
    )
    cartera_parser.add_argument(
        "--output",
        required=True,
        help="Ruta base del JSON de salida. Se generarán CSVs auxiliares al lado.",
    )

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    try:
        config = build_config()

        if args.command == "auth-test":
            token, token_type = obtain_token(config, timeout=args.timeout)
            print("Autenticación exitosa.")
            print(f"Tipo de token: {token_type}")
            print(f"Token enmascarado: {mask_token(token)}")
            return 0

        if args.command == "get":
            params = parse_params(args.param)
            status, body = readonly_get(config, args.path, params, timeout=args.timeout)
            print(f"GET {normalize_path(args.path)} -> HTTP {status}")
            print_body(body)

            if args.output:
                output_path = Path(args.output)
                write_output(output_path, body)
                print(f"\nRespuesta guardada en: {output_path}")
            return 0

        if args.command == "export":
            params = parse_params(args.param)
            if args.max_pages is not None and args.max_pages < 1:
                raise SiigoApiError("--max-pages debe ser mayor o igual a 1")

            body = export_paginated_get(
                config,
                args.path,
                params,
                timeout=args.timeout,
                max_pages=args.max_pages,
            )
            write_output(Path(args.output), body)

            summary = body["summary"]
            print(f"EXPORT {summary['path']} -> páginas: {summary['pages_fetched']}")
            print(f"Registros exportados: {summary['exported_results']}")
            print(f"Total reportado por Siigo: {summary['reported_total_results']}")
            print(f"Archivo generado: {args.output}")
            return 0

        if args.command == "sales-report":
            if args.months < 1:
                raise SiigoApiError("--months debe ser mayor o igual a 1")
            if args.page_size < 1:
                raise SiigoApiError("--page-size debe ser mayor o igual a 1")
            if args.max_pages is not None and args.max_pages < 1:
                raise SiigoApiError("--max-pages debe ser mayor o igual a 1")

            report = build_sales_report(
                config,
                months=args.months,
                from_date_raw=args.from_date,
                to_date_raw=args.to_date,
                include_annulled=args.include_annulled,
                page_size=args.page_size,
                timeout=args.timeout,
                max_pages=args.max_pages,
            )
            output_files = write_sales_report_files(Path(args.output), report)
            summary = report["summary"]

            print(
                f"SALES REPORT {summary['from_date']} -> {summary['to_date']} "
                f"| facturas: {summary['report_invoices']}"
            )
            print(f"Clientes: {summary['customers']}")
            print(f"Ventas totales: {summary['total_sales']}")
            print(f"Saldo pendiente: {summary['outstanding_balance']}")
            print(f"Páginas consultadas: {summary['pages_fetched']}")
            print(f"JSON: {output_files['json']}")
            print(f"XLSX: {output_files['xlsx']}")
            print(f"CSV clientes: {output_files['by_customer_csv']}")
            print(f"CSV cliente-mes: {output_files['by_customer_month_csv']}")
            print(f"CSV facturas: {output_files['invoices_csv']}")
            return 0

        if args.command == "billing-report":
            if args.months < 1:
                raise SiigoApiError("--months debe ser mayor o igual a 1")
            if args.page_size < 1:
                raise SiigoApiError("--page-size debe ser mayor o igual a 1")
            if args.max_pages is not None and args.max_pages < 1:
                raise SiigoApiError("--max-pages debe ser mayor o igual a 1")

            report = build_billing_report(
                config,
                months=args.months,
                from_date_raw=args.from_date,
                to_date_raw=args.to_date,
                include_annulled=args.include_annulled,
                page_size=args.page_size,
                timeout=args.timeout,
                max_pages=args.max_pages,
            )
            output_files = write_billing_report_files(Path(args.output), report)
            summary = report["summary"]

            print(
                f"BILLING REPORT {summary['from_date']} -> {summary['to_date']} "
                f"| facturas: {summary['report_invoices']}"
            )
            print(f"Clientes: {summary['customers']}")
            print(f"Vendedores: {summary['sellers']}")
            print(f"Cliente-vendedor: {summary['customer_seller_pairs']}")
            print(f"Ventas totales: {summary['total_sales']}")
            print(f"Saldo pendiente: {summary['outstanding_balance']}")
            print(f"Páginas consultadas: {summary['pages_fetched']}")
            print(f"JSON: {output_files['json']}")
            print(f"XLSX: {output_files['xlsx']}")
            print(f"CSV clientes: {output_files['by_customer_csv']}")
            print(f"CSV vendedores: {output_files['by_seller_csv']}")
            print(f"CSV cliente-vendedor: {output_files['by_customer_seller_csv']}")
            print(f"CSV por día: {output_files['by_day_csv']}")
            print(f"CSV facturas: {output_files['invoices_csv']}")
            return 0

        if args.command == "cartera-report":
            if args.months < 1:
                raise SiigoApiError("--months debe ser mayor o igual a 1")
            if args.page_size < 1:
                raise SiigoApiError("--page-size debe ser mayor o igual a 1")
            if args.max_pages is not None and args.max_pages < 1:
                raise SiigoApiError("--max-pages debe ser mayor o igual a 1")

            report = build_cartera_report(
                config,
                months=args.months,
                from_date_raw=args.from_date,
                to_date_raw=args.to_date,
                page_size=args.page_size,
                timeout=args.timeout,
                max_pages=args.max_pages,
            )
            output_files = write_cartera_report_files(Path(args.output), report)
            summary = report["summary"]

            print(
                f"CARTERA REPORT {summary['from_date']} -> {summary['to_date']} "
                f"| facturas con saldo: {summary['invoices_with_balance']}"
            )
            print(f"Clientes con deuda: {summary['customers_with_debt']}")
            print(f"Saldo total pendiente: {summary['total_balance']}")
            print(f"Páginas consultadas: {summary['pages_fetched']}")
            print(f"JSON: {output_files['json']}")
            print(f"XLSX: {output_files['xlsx']}")
            print(f"CSV clientes: {output_files['by_customer_csv']}")
            print(f"CSV facturas: {output_files['invoices_csv']}")
            return 0

        parser.print_help()
        return 1
    except SiigoApiError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
